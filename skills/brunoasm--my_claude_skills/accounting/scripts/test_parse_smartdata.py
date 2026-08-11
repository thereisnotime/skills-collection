#!/usr/bin/env python3
"""
Self-tests for parse_smartdata.py.

There is no pytest in this repo, so run this file directly:

    python3 accounting/scripts/test_parse_smartdata.py

Fixtures come from _smartdata_fixtures.py and are entirely synthetic, so no
real financial data or cardholder PII is ever committed.
"""

import sys
import tempfile
from decimal import Decimal
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _smartdata_fixtures import build_workbook, pad, txn  # noqa: E402
from parse_smartdata import (  # noqa: E402
    _clean,
    _number,
    fee_of,
    find_header_row,
    is_international,
    parse_report,
    parse_transactions,
    statement_period,
)


def load(tmpdir, **kwargs):
    path = build_workbook(Path(tmpdir) / "report.xlsx", **kwargs)
    wb = openpyxl.load_workbook(path)
    return path, wb


def test_clean_collapses_newlines_and_blanks():
    assert _clean("\nPosting Date") == "Posting Date"
    assert _clean("Transaction\nDate") == "Transaction Date"
    assert _clean(" ") == ""
    assert _clean(None) == ""


def test_number_strips_padding_and_separators():
    assert _number(pad("31.99")) == Decimal("31.99")
    assert _number("1,234.56") == Decimal("1234.56")
    assert _number(" ") == Decimal("0")


def test_header_row_found_below_metadata():
    with tempfile.TemporaryDirectory() as tmp:
        _, wb = load(tmp)
        assert find_header_row(wb["Detail Report"]) == 14


def test_addendum_rows_are_skipped():
    with tempfile.TemporaryDirectory() as tmp:
        _, wb = load(tmp)
        txns = parse_transactions(wb["Detail Report"])
        # 3 purchases + 2 fee lines; the 2 addendum rows must not appear
        assert len(txns) == 5
        assert all(t["description"] != "Quantity:" for t in txns)


def test_padded_numbers_become_decimals():
    with tempfile.TemporaryDirectory() as tmp:
        _, wb = load(tmp)
        first = parse_transactions(wb["Detail Report"])[0]
        assert first["amount"] == Decimal("25.00")
        assert first["original_currency"] == "USD"
        assert first["country"] == "UNITED STATES"


def test_fee_rows_have_blank_country():
    with tempfile.TemporaryDirectory() as tmp:
        _, wb = load(tmp)
        fees = [t for t in parse_transactions(wb["Detail Report"])
                if t["description"] == "INTERNATIONAL TRANSACTION"]
        assert len(fees) == 2
        assert all(f["country"] == "" for f in fees)


def test_statement_period_read_from_metadata():
    with tempfile.TemporaryDirectory() as tmp:
        _, wb = load(tmp)
        assert statement_period(wb["Detail Report"]) == "01/26/2026 - 02/25/2026"


def test_fee_of_rounds_half_up():
    # 12.50 * 1% == 0.125 exactly -- half-up gives 0.13, bankers' rounding 0.12
    assert fee_of(Decimal("12.50")) == Decimal("0.13")
    assert fee_of(Decimal("29.99")) == Decimal("0.30")
    assert fee_of(Decimal("3.00")) == Decimal("0.03")


def test_is_international_uses_country():
    assert is_international({"country": "CANADA"}) is True
    assert is_international({"country": "KOREA, REPUBLIC OF"}) is True
    assert is_international({"country": "UNITED STATES"}) is False
    assert is_international({"country": ""}) is False


def test_pairs_each_fee_with_its_purchase():
    with tempfile.TemporaryDirectory() as tmp:
        path, _ = load(tmp)
        result = parse_report(path)
        assert len(result["purchases"]) == 3
        assert len(result["fees"]) == 2
        assert len(result["pairs"]) == 2
        assert result["problems"] == []
        assert result["international_without_fee"] == []
        paired = {p["purchase"]["description"]: p["fee"]["amount"] for p in result["pairs"]}
        assert paired["EXAMPLE CA VENDOR"] == Decimal("0.13")
        assert paired["EXAMPLE PH VENDOR"] == Decimal("0.30")


def test_validation_passes_against_summary():
    with tempfile.TemporaryDirectory() as tmp:
        path, _ = load(tmp)
        validation = parse_report(path)["validation"]
        assert validation["ok"] is True
        assert all(c["ok"] for c in validation["checks"])


def test_validation_flags_summary_mismatch():
    with tempfile.TemporaryDirectory() as tmp:
        path = build_workbook(Path(tmp) / "report.xlsx",
                              summary_row=[99, "1.00", 99, "1.00", 198, "2.00"])
        validation = parse_report(path)["validation"]
        assert validation["ok"] is False
        failed = {c["name"] for c in validation["checks"] if not c["ok"]}
        assert failed == {"purchase_count", "purchase_total", "fee_count", "fee_total"}


def test_ambiguous_pairing_is_reported_not_guessed():
    rows = [
        txn("02/03/2026", "02/02/2026", "EXAMPLE CA VENDOR A", "TORONTO, ON",
            "CANADA", "17.50", "CAD", "1.4000", "12.50"),
        txn("02/03/2026", "02/02/2026", "EXAMPLE CA VENDOR B", "TORONTO, ON",
            "CANADA", "17.50", "CAD", "1.4000", "12.50"),
        txn("02/03/2026", "02/02/2026", "INTERNATIONAL TRANSACTION", " ", " ",
            "0.13", "USD", "1.0000", "0.13"),
    ]
    with tempfile.TemporaryDirectory() as tmp:
        path = build_workbook(Path(tmp) / "report.xlsx", txn_rows=rows,
                              summary_row=[2, "25.00", 1, "0.13", 3, "25.13"])
        result = parse_report(path)
        assert result["pairs"] == []
        assert len(result["problems"]) == 1
        problem = result["problems"][0]
        assert problem["reason"] == "ambiguous"
        assert len(problem["candidates"]) == 2
        # both candidates cost the same, so either assignment gives the same numbers
        assert problem["equivalent"] is True
        # ambiguous candidates must not appear in international_without_fee
        assert result["international_without_fee"] == []


def test_fee_with_no_candidate_is_reported():
    rows = [
        txn("02/03/2026", "02/02/2026", "EXAMPLE US VENDOR", "CHICAGO, IL",
            "UNITED STATES", "25.00", "USD", "1.0000", "25.00"),
        txn("02/03/2026", "02/02/2026", "INTERNATIONAL TRANSACTION", " ", " ",
            "0.99", "USD", "1.0000", "0.99"),
    ]
    with tempfile.TemporaryDirectory() as tmp:
        path = build_workbook(Path(tmp) / "report.xlsx", txn_rows=rows,
                              summary_row=[1, "25.00", 1, "0.99", 2, "25.99"])
        result = parse_report(path)
        assert [p["reason"] for p in result["problems"]] == ["no_match"]
        assert result["problems"][0]["candidates"] == []


def test_international_purchase_without_fee_is_reported():
    rows = [
        txn("02/03/2026", "02/02/2026", "EXAMPLE CA VENDOR", "TORONTO, ON",
            "CANADA", "17.50", "CAD", "1.4000", "12.50"),
    ]
    with tempfile.TemporaryDirectory() as tmp:
        path = build_workbook(Path(tmp) / "report.xlsx", txn_rows=rows,
                              summary_row=[1, "12.50", 0, "0.00", 1, "12.50"])
        result = parse_report(path)
        assert [p["description"] for p in result["international_without_fee"]] \
            == ["EXAMPLE CA VENDOR"]


def test_output_carries_no_cardholder_pii():
    import json
    with tempfile.TemporaryDirectory() as tmp:
        path, _ = load(tmp)
        dumped = json.dumps(parse_report(path), default=str)
        for secret in ("XX -00000000", "TAX EX", "EXAMPLE ST", "TEST CARDHOLDER"):
            assert secret not in dumped


def test_summary_with_unexpected_header_falls_back_instead_of_raising():
    """A Summary sheet laid out differently must degrade, not crash the parse.

    Without the fallback, find_header_row's ValueError escapes parse_report and
    validate()'s summary_present branch is unreachable.
    """
    with tempfile.TemporaryDirectory() as tmp:
        path = build_workbook(Path(tmp) / "report.xlsx")
        wb = openpyxl.load_workbook(path)
        summary = wb["Summary Report"]
        header = find_header_row(summary, "Account Name")
        summary.cell(row=header, column=1).value = "Cardholder"
        wb.save(path)

        result = parse_report(path)
        # the detail sheet still parses
        assert len(result["purchases"]) == 3
        validation = result["validation"]
        assert validation["ok"] is False
        assert [c["name"] for c in validation["checks"]] == ["summary_present"]
        assert validation["checks"][0]["ok"] is False


TESTS = [v for k, v in sorted(globals().items()) if k.startswith("test_")]


def main():
    failures = []
    for test in TESTS:
        try:
            test()
            print(f"  ok    {test.__name__}")
        except AssertionError as exc:
            failures.append(test.__name__)
            print(f"  FAIL  {test.__name__}: {exc or 'assertion failed'}")
        except Exception as exc:  # noqa: BLE001
            failures.append(test.__name__)
            print(f"  ERROR {test.__name__}: {type(exc).__name__}: {exc}")
    print(f"\n{len(TESTS) - len(failures)}/{len(TESTS)} passed")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
