#!/usr/bin/env python3
"""
Parses a J.P. Morgan SmartData "Account Statement (Version 2)" XLSX export.

Layout, verified against a real statement:

  Sheet "Detail Report"
    rows 1-12   report metadata, then the cardholder's name, tax id, card
                number and street address. This parser never emits those.
    row ~13     blank
    row ~14     table header. Cells contain embedded newlines, e.g.
                "Transaction\\nDate". Found by locating the row whose first
                cell cleans to "Posting Date" -- never hard-code the number.
    rows ~15+   Posting Date | Transaction Date | Description | Location |
                Country | Original Amount | Original Currency Code |
                Conversion Rate | Amount

Traps this module exists to absorb:

  * A row is a transaction only if its first cell looks like MM/DD/YYYY. Every
    other row is Level-3 addendum detail belonging to the transaction above it
    ("Description:", "Quantity:", "Guest Name:", "Total Room Nights:").
  * Numeric cells are whitespace-padded strings; summary figures also carry
    thousands separators.
  * Fee rows have Description == "INTERNATIONAL TRANSACTION" with blank
    Location and Country.
"""

import argparse
import json
import re
import sys
import warnings
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import openpyxl

FEE_DESCRIPTION = "INTERNATIONAL TRANSACTION"
FEE_RATE = Decimal("0.01")
DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
DETAIL_SHEET = "Detail Report"
SUMMARY_SHEET = "Summary Report"

_COLUMN_COUNT = 9


def _clean(value):
    """Reduce a cell to comparable text: no newlines, collapsed whitespace."""
    if value is None:
        return ""
    return " ".join(str(value).split())


def _number(value):
    """Parse a whitespace-padded, comma-separated numeric cell."""
    text = _clean(value).replace(",", "")
    if not text:
        return Decimal("0")
    return Decimal(text)


def find_header_row(ws, first_header="Posting Date"):
    """Return the 1-based row index of the table header."""
    for row in ws.iter_rows(min_row=1, max_row=40):
        if row and _clean(row[0].value) == first_header:
            return row[0].row
    raise ValueError(
        f'header row starting with "{first_header}" not found in sheet "{ws.title}"'
    )


def statement_period(ws):
    """Return the posting-date window from the metadata block, e.g. "A - B"."""
    for row in ws.iter_rows(min_row=1, max_row=13, values_only=True):
        text = _clean(row[0]) if row else ""
        if text.startswith("Posting Date:"):
            return text.split(":", 1)[1].strip()
    return ""


def parse_transactions(ws):
    """Return the transaction rows, skipping Level-3 addendum detail."""
    header_row = find_header_row(ws)
    transactions = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        cells = list(row) + [None] * (_COLUMN_COUNT - len(row))
        if not DATE_RE.match(_clean(cells[0])):
            continue
        transactions.append({
            "posting_date": _clean(cells[0]),
            "transaction_date": _clean(cells[1]),
            "description": _clean(cells[2]),
            "location": _clean(cells[3]),
            "country": _clean(cells[4]),
            "original_amount": _number(cells[5]),
            "original_currency": _clean(cells[6]),
            "conversion_rate": _number(cells[7]),
            "amount": _number(cells[8]),
        })
    return transactions


def load_workbook(path):
    """Open a report, silencing openpyxl's missing-default-style warning."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return openpyxl.load_workbook(path, data_only=True)


def fee_of(amount):
    """The international transaction fee on a posted USD amount: 1%, half-up."""
    return (amount * FEE_RATE).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def is_international(txn):
    """True when the report says the charge was processed outside the US.

    Fee rows carry a blank Country, so they are not themselves international
    purchases.
    """
    return txn["country"] not in ("", "UNITED STATES")


def parse_summary(ws):
    """Return the "Report Totals" figures, or None when absent."""
    header_row = find_header_row(ws, "Account Name")
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        if _clean(row[0]) == "Report Totals":
            cells = list(row) + [None] * (7 - len(row))
            return {
                "transaction_count": int(_number(cells[1])),
                "transaction_amount": _number(cells[2]),
                "payment_count": int(_number(cells[3])),
                "payment_amount": _number(cells[4]),
                "total_count": int(_number(cells[5])),
                "total_amount": _number(cells[6]),
            }
    return None


def pair_fees(purchases, fees):
    """Attribute each fee line to the purchase that incurred it.

    Fee rows carry no merchant name, so a fee pairs with a purchase when their
    transaction dates match, the purchase is international, and 1% of the
    purchase's USD amount equals the fee. Anything other than exactly one
    candidate is returned as a problem for a human to resolve -- never guessed.
    """
    unpaired = list(purchases)
    pairs, problems = [], []
    for fee in fees:
        candidates = [
            p for p in unpaired
            if p["transaction_date"] == fee["transaction_date"]
            and is_international(p)
            and fee_of(p["amount"]) == fee["amount"]
        ]
        if len(candidates) == 1:
            unpaired.remove(candidates[0])
            pairs.append({"fee": fee, "purchase": candidates[0]})
        else:
            problems.append({
                "fee": fee,
                "reason": "ambiguous" if candidates else "no_match",
                "candidates": candidates,
                "equivalent": bool(candidates)
                and len({c["amount"] for c in candidates}) == 1,
            })
    return pairs, problems, unpaired


def validate(purchases, fees, summary):
    """Cross-check the parse against the statement's own summary totals."""
    if summary is None:
        return {"ok": False, "checks": [{
            "name": "summary_present", "expected": "Report Totals row",
            "actual": "not found", "ok": False,
            "note": "counts and totals could not be cross-checked",
        }]}

    purchase_total = sum((p["amount"] for p in purchases), Decimal("0"))
    fee_total = sum((f["amount"] for f in fees), Decimal("0"))
    payment_note = (
        "real payments and refunds also land in the statement's Payment bucket, "
        "so a mismatch here means review, not a broken parse"
    )
    checks = [
        {"name": "purchase_count", "expected": summary["transaction_count"],
         "actual": len(purchases),
         "ok": len(purchases) == summary["transaction_count"],
         "note": "purchases parsed vs statement Transaction Count"},
        {"name": "purchase_total", "expected": str(summary["transaction_amount"]),
         "actual": str(purchase_total),
         "ok": purchase_total == summary["transaction_amount"],
         "note": "purchase sum vs statement Transaction Amount"},
        {"name": "fee_count", "expected": summary["payment_count"],
         "actual": len(fees), "ok": len(fees) == summary["payment_count"],
         "note": payment_note},
        {"name": "fee_total", "expected": str(summary["payment_amount"]),
         "actual": str(fee_total), "ok": fee_total == summary["payment_amount"],
         "note": payment_note},
    ]
    return {"ok": all(c["ok"] for c in checks), "checks": checks}


def parse_report(path):
    """Parse a report and pair its fees. Never returns cardholder PII."""
    wb = load_workbook(path)
    detail = wb[DETAIL_SHEET]
    transactions = parse_transactions(detail)

    # A Summary sheet that is present but laid out differently must degrade to
    # the same "could not cross-check" verdict as a missing one -- validate()'s
    # summary_present branch -- rather than crashing the whole parse.
    summary = None
    if SUMMARY_SHEET in wb.sheetnames:
        try:
            summary = parse_summary(wb[SUMMARY_SHEET])
        except ValueError:
            summary = None

    fees = [t for t in transactions if t["description"] == FEE_DESCRIPTION]
    purchases = [t for t in transactions if t["description"] != FEE_DESCRIPTION]
    pairs, problems, unpaired = pair_fees(purchases, fees)

    # Exclude purchases that are referenced in ambiguous problems from international_without_fee
    ambiguous_candidates = {id(p) for problem in problems for p in problem["candidates"]}
    truly_unpaired = [p for p in unpaired if id(p) not in ambiguous_candidates]

    return {
        "period": statement_period(detail),
        "purchases": purchases,
        "fees": fees,
        "pairs": pairs,
        "problems": problems,
        "international_without_fee": [p for p in truly_unpaired if is_international(p)],
        "validation": validate(purchases, fees, summary),
    }


def main():
    parser = argparse.ArgumentParser(
        description="Parse a SmartData Account Statement XLSX and pair "
                    "international transaction fees with their purchases.")
    parser.add_argument("report", type=Path, help="path to a YYYY_MM.xlsx report")
    parser.add_argument("--json", action="store_true",
                        help="emit JSON instead of a human-readable summary")
    args = parser.parse_args()

    result = parse_report(args.report)

    if args.json:
        print(json.dumps(result, indent=2, default=str))
        return 0 if result["validation"]["ok"] else 1

    print(f"statement period: {result['period'] or 'unknown'}")
    print(f"purchases: {len(result['purchases'])}   fee lines: {len(result['fees'])}")
    print("validation:")
    for check in result["validation"]["checks"]:
        flag = "ok" if check["ok"] else "REVIEW"
        print(f"  [{flag}] {check['name']}: expected {check['expected']}, "
              f"got {check['actual']}")
    if result["pairs"]:
        print("paired fees:")
        for pair in result["pairs"]:
            p, f = pair["purchase"], pair["fee"]
            print(f"  {p['description']} ({p['country']}) "
                  f"{p['original_amount']} {p['original_currency']} "
                  f"-> ${p['amount']}  fee ${f['amount']}")
    for problem in result["problems"]:
        fee = problem["fee"]
        print(f"  ASK: fee ${fee['amount']} on {fee['transaction_date']} "
              f"-> {problem['reason']} ({len(problem['candidates'])} candidates"
              f"{', equivalent' if problem['equivalent'] else ''})")
    for purchase in result["international_without_fee"]:
        print(f"  ASK: international purchase with no fee line: "
              f"{purchase['description']} ${purchase['amount']}")
    return 0 if result["validation"]["ok"] else 1


if __name__ == "__main__":
    sys.exit(main())
