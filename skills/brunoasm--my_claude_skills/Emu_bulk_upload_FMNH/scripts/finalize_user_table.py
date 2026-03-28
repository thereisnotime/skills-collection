#!/usr/bin/env python3
"""
Add ColSiteRef.irn column to the original user xlsx file.

Usage:
    python3 finalize_user_table.py <original.xlsx> <irn_mapping.json> <output.xlsx>

irn_mapping.json should be a JSON file with format:
    {"row_irn_map": {"4": "123456", "5": "123457", ...}}
    where keys are original xlsx row numbers and values are site IRNs.
"""

import json
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Error: openpyxl is required. Install with: pip install openpyxl")

GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")


def find_site_number_column(ws):
    """Find the column index of SitSiteNumber in row 2."""
    for cell in ws[2]:
        if cell.value and str(cell.value).strip() == "SitSiteNumber":
            return cell.column
    return None


def insert_irn_column(ws, after_col, irn_map, data_start_row=4):
    """Insert ColSiteRef.irn column after the specified column."""
    insert_col = after_col + 1

    # Insert column (shifts everything right)
    ws.insert_cols(insert_col)

    # Row 1: user-friendly name
    cell1 = ws.cell(row=1, column=insert_col, value="Site IRN")
    cell1.fill = GREEN_FILL

    # Row 2: Emu field name
    cell2 = ws.cell(row=2, column=insert_col, value="ColSiteRef.irn")
    cell2.fill = GREEN_FILL

    # Row 3: example (leave blank or add note)
    cell3 = ws.cell(row=3, column=insert_col, value="")
    cell3.fill = GREEN_FILL

    # Data rows
    filled = 0
    for row_str, irn in irn_map.items():
        row_num = int(row_str)
        if row_num >= data_start_row:
            cell = ws.cell(row=row_num, column=insert_col, value=irn)
            cell.fill = GREEN_FILL
            filled += 1

    return insert_col, filled


def main():
    if len(sys.argv) < 4:
        print(f"Usage: {sys.argv[0]} <original.xlsx> <irn_mapping.json> <output.xlsx>")
        sys.exit(1)

    original_path = sys.argv[1]
    irn_path = sys.argv[2]
    output_path = sys.argv[3]

    # Load IRN mapping
    with open(irn_path) as f:
        irn_data = json.load(f)

    irn_map = irn_data.get("row_irn_map", {})
    if not irn_map:
        sys.exit("Error: irn_mapping.json has no row_irn_map entries")

    # Load workbook (preserve formatting)
    wb = openpyxl.load_workbook(original_path)
    ws = wb.active

    # Find SitSiteNumber column
    site_num_col = find_site_number_column(ws)
    if site_num_col is None:
        print("Warning: SitSiteNumber column not found in row 2. "
              "Adding ColSiteRef.irn at the end.")
        site_num_col = ws.max_column

    insert_col, filled = insert_irn_column(ws, site_num_col, irn_map)

    print(f"Inserted ColSiteRef.irn at column {get_column_letter(insert_col)} "
          f"(after SitSiteNumber at {get_column_letter(site_num_col)})")
    print(f"Filled {filled} IRN values out of {len(irn_map)} in mapping")

    wb.save(output_path)
    print(f"Saved to {output_path}")


if __name__ == "__main__":
    main()
