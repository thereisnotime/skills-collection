#!/usr/bin/env python3
"""
Parse a user xlsx file and extract site columns (green fill).

Usage:
    python3 parse_user_data.py <input.xlsx> [output.json]

Outputs JSON with:
    - site_columns: list of {col_index, col_letter, user_name, emu_name, canonical_name}
    - site_records: list of dicts (one per data row) with canonical field names
    - row_mapping: maps each site record index to the original xlsx row number
    - metadata: column range, data_start_row, total_rows
"""

import json
import sys
import os

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Error: openpyxl is required. Install with: pip install openpyxl")


# Green fill codes that indicate site module columns
GREEN_FILLS = {"FFCCFFCC", "00CCFFCC", "CCFFCC"}

# Site hierarchy fields in order (most general to most precise)
HIERARCHY_FIELDS = [
    "LocContinent",
    "LocCountry",
    "LocProvinceStateTerritory",
    "LocDistrictCountyShire",
    "LocTownship",
    "LocPreciseLocation",
]

# All expected site fields (hierarchy + elevation + coordinates + site number)
SITE_FIELDS = HIERARCHY_FIELDS + [
    "LocElevationASLFromMt",
    "LocElevationASLToMt",
    "LocElevationASLFromFt",
    "LocElevationASLToFt",
    "LatLatitude",
    "LatLongitude",
    "SitSiteNumber",
]


def normalize_field_name(emu_name):
    """Strip _tab, _nesttab suffixes to get canonical Emu field name."""
    if not emu_name:
        return ""
    name = str(emu_name).strip()
    for suffix in ("_nesttab", "_tab"):
        if name.endswith(suffix):
            name = name[: -len(suffix)]
    return name


def is_green_fill(cell):
    """Check if a cell has green fill indicating site module."""
    fill = cell.fill
    if fill and fill.fgColor:
        rgb = fill.fgColor.rgb
        if isinstance(rgb, str):
            # Strip leading "FF" transparency prefix if 8 chars
            color = rgb.upper()
            return color in GREEN_FILLS or color[2:] in GREEN_FILLS
    return False


def identify_site_columns(ws):
    """Find site columns by green fill in row 1 or by matching known field names in row 2."""
    site_cols = []

    for cell in ws[1]:  # Row 1: user-friendly names
        col_idx = cell.column
        emu_name = ws.cell(row=2, column=col_idx).value
        canonical = normalize_field_name(emu_name)

        # Check by green fill
        green = is_green_fill(cell) or is_green_fill(ws.cell(row=2, column=col_idx))

        # Check by known field name
        known = canonical in SITE_FIELDS

        if green or known:
            # Skip columns with no Emu field name (e.g., UTM columns without Row 2 labels)
            if not canonical:
                continue
            site_cols.append(
                {
                    "col_index": col_idx,
                    "col_letter": get_column_letter(col_idx),
                    "user_name": str(cell.value or ""),
                    "emu_name": str(emu_name or ""),
                    "canonical_name": canonical,
                }
            )

    return site_cols


def extract_site_data(ws, site_cols, data_start_row=4):
    """Extract site data from each data row."""
    records = []
    row_mapping = []

    for row_idx in range(data_start_row, ws.max_row + 1):
        # Skip fully empty rows
        has_data = False
        record = {}
        for col_info in site_cols:
            val = ws.cell(row=row_idx, column=col_info["col_index"]).value
            if val is not None and str(val).strip() != "":
                has_data = True
                # Convert floats that are really ints
                if isinstance(val, float) and val == int(val) and col_info["canonical_name"] not in ("LatLatitude", "LatLongitude"):
                    val = int(val)
                record[col_info["canonical_name"]] = val
            else:
                record[col_info["canonical_name"]] = None

        if has_data:
            records.append(record)
            row_mapping.append(row_idx)

    return records, row_mapping


def main():
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <input.xlsx> [output.json]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active

    site_cols = identify_site_columns(ws)
    if not site_cols:
        sys.exit("Error: No site columns found (no green fill or known field names)")

    records, row_mapping = extract_site_data(ws, site_cols)

    result = {
        "source_file": os.path.basename(input_path),
        "metadata": {
            "total_data_rows": len(records),
            "data_start_row": 4,
            "site_column_range": f"{site_cols[0]['col_letter']}-{site_cols[-1]['col_letter']}",
            "site_column_count": len(site_cols),
        },
        "site_columns": site_cols,
        "site_records": records,
        "row_mapping": row_mapping,
    }

    output = json.dumps(result, indent=2, default=str)

    if output_path:
        with open(output_path, "w") as f:
            f.write(output)
        print(f"Wrote {len(records)} site records to {output_path}")
    else:
        print(output)


if __name__ == "__main__":
    main()
