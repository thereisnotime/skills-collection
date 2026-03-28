#!/usr/bin/env python3
"""
Generate Emu bulk upload xlsx tables for new site records.

Usage:
    python3 generate_bulk_upload.py <parent_results.json> <output_dir>

Creates one or more xlsx files for uploading new sites to Emu.
Each file is a batch — if sites depend on parents that also need creation,
multiple batches are generated (highest-level parents first).

Bulk upload columns:
    - Site hierarchy fields (only those NOT implied by the parent)
    - Elevation fields
    - Coordinate fields
    - PolParent (parent IRN)
"""

import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    sys.exit("Error: openpyxl is required. Install with: pip install openpyxl")

# Green fill for site columns
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

HIERARCHY_FIELDS = [
    "LocContinent",
    "LocCountry",
    "LocProvinceStateTerritory",
    "LocDistrictCountyShire",
    "LocTownship",
    "LocPreciseLocation",
]

# Map from hierarchy field to the level index at which a parent covers it
HIERARCHY_INDEX = {f: i for i, f in enumerate(HIERARCHY_FIELDS)}

# All site columns in upload order
UPLOAD_COLUMNS = [
    ("LocContinent", "Continent"),
    ("LocCountry", "Country"),
    ("LocProvinceStateTerritory", "Province/State"),
    ("LocDistrictCountyShire", "County"),
    ("LocTownship", "City"),
    ("LocPreciseLocation", "Precise Location"),
    ("LocElevationASLFromMt", "Elevation From Mt."),
    ("LocElevationASLToMt", "Elevation To Mt."),
    ("LocElevationASLFromFt", "Elevation From Ft."),
    ("LocElevationASLToFt", "Elevation To Ft."),
    ("LatLatitude", "Latitude"),
    ("LatLongitude", "Longitude"),
    ("PolParent", "Parent IRN"),
]


def get_parent_level(parent_search):
    """Determine the hierarchy level of the parent."""
    if not parent_search:
        return -1
    search_level = parent_search.get("search_level", "")
    return HIERARCHY_INDEX.get(search_level, -1)


def build_upload_row(user_site, parent_irn, parent_level):
    """Build a row for the bulk upload table.
    Fields at the parent's level and above are left blank
    (they are implied by the parent).
    """
    row = {}
    for emu_field, _ in UPLOAD_COLUMNS:
        if emu_field == "PolParent":
            row[emu_field] = parent_irn
        elif emu_field in HIERARCHY_INDEX:
            # Only include fields BELOW the parent level
            field_level = HIERARCHY_INDEX[emu_field]
            if field_level > parent_level:
                row[emu_field] = user_site.get(emu_field)
            else:
                row[emu_field] = None  # Implied by parent
        else:
            row[emu_field] = user_site.get(emu_field)
    return row


def deduplicate_upload_rows(rows_with_meta):
    """Deduplicate upload rows (same site data + same parent = same record)."""
    seen = {}
    unique = []
    for entry in rows_with_meta:
        row = entry["row"]
        key = tuple(str(row.get(f, "")) for f, _ in UPLOAD_COLUMNS)
        if key not in seen:
            seen[key] = len(unique)
            unique.append(entry)
        else:
            # Merge original rows
            idx = seen[key]
            unique[idx]["original_site_indices"].extend(entry["original_site_indices"])
    return unique


def write_upload_xlsx(rows_with_meta, output_path, batch_num):
    """Write a bulk upload xlsx file."""
    # Filter out columns that are entirely blank across all data rows
    active_columns = []
    for emu_field, user_name in UPLOAD_COLUMNS:
        has_data = any(
            entry["row"].get(emu_field) is not None and entry["row"].get(emu_field) != ""
            for entry in rows_with_meta
        )
        if has_data:
            active_columns.append((emu_field, user_name))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sites Batch {batch_num}"

    # Row 1: user-friendly names
    for col_idx, (emu_field, user_name) in enumerate(active_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=user_name)
        cell.fill = GREEN_FILL

    # Row 2: Emu field names
    for col_idx, (emu_field, _) in enumerate(active_columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=emu_field)
        cell.fill = GREEN_FILL

    # Data rows starting at row 3
    for row_idx, entry in enumerate(rows_with_meta, 3):
        row = entry["row"]
        for col_idx, (emu_field, _) in enumerate(active_columns, 1):
            val = row.get(emu_field)
            if val is not None:
                ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-adjust column widths
    for col_idx in range(1, len(active_columns) + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 30)

    if active_columns != [(f, n) for f, n in UPLOAD_COLUMNS]:
        removed = set(f for f, _ in UPLOAD_COLUMNS) - set(f for f, _ in active_columns)
        if removed:
            print(f"  Removed {len(removed)} empty column(s): {', '.join(sorted(removed))}")

    wb.save(output_path)
    return len(rows_with_meta)


def generate_batches(parent_results):
    """Organize sites into upload batches.
    Batch 1: sites whose parents already exist in Emu
    Batch 2+: sites whose parents need to be created first (future)
    """
    batch1_rows = []

    for result in parent_results["results"]:
        if result["match_status"] == "already_matched":
            continue  # Already has an IRN, no upload needed

        parent = result.get("parent_search")
        if not parent:
            continue

        if parent["status"] in ("perfect", "partial") and parent.get("parent_irn"):
            parent_irn = parent["parent_irn"]
            parent_level = get_parent_level(parent)
            user_site = result["user_site"]

            row = build_upload_row(user_site, parent_irn, parent_level)
            batch1_rows.append({
                "row": row,
                "site_index": result["site_index"],
                "original_site_indices": [result["site_index"]],
                "location_label": result["location_label"],
                "parent_irn": parent_irn,
                "parent_level": HIERARCHY_FIELDS[parent_level] if parent_level >= 0 else "unknown",
            })
        elif parent["status"] == "not_found":
            # These need parent creation first — future batch
            print(f"  Warning: Site {result['site_index']} ({result['location_label']}) "
                  f"needs parent creation (not implemented in this batch)")

    return [batch1_rows] if batch1_rows else []


def main():
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} <parent_results.json> <output_dir>")
        sys.exit(1)

    parent_path = sys.argv[1]
    output_dir = sys.argv[2]

    with open(parent_path) as f:
        parent_data = json.load(f)

    os.makedirs(output_dir, exist_ok=True)

    print("Generating bulk upload tables...")
    batches = generate_batches(parent_data)

    if not batches:
        print("No new sites to upload.")
        return

    for batch_num, batch_rows in enumerate(batches, 1):
        # Deduplicate
        unique_rows = deduplicate_upload_rows(batch_rows)

        output_path = os.path.join(output_dir, f"sites_upload_batch_{batch_num}.xlsx")
        count = write_upload_xlsx(unique_rows, output_path, batch_num)
        print(f"  Batch {batch_num}: {count} unique sites → {output_path}")

    # Also save metadata for tracking
    meta = {
        "batches": [
            {
                "batch": i + 1,
                "site_count": len(deduplicate_upload_rows(batch)),
                "sites": [
                    {
                        "site_index": r["site_index"],
                        "location": r["location_label"],
                        "parent_irn": r["parent_irn"],
                    }
                    for r in batch
                ],
            }
            for i, batch in enumerate(batches)
        ]
    }
    meta_path = os.path.join(output_dir, "upload_metadata.json")
    with open(meta_path, "w") as f:
        json.dump(meta, f, indent=2)
    print(f"  Metadata → {meta_path}")


if __name__ == "__main__":
    main()
