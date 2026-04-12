#!/usr/bin/env python3
"""
多格式导出模块 - 将微信文章导出为多种格式

支持格式：
- Markdown (默认)
- PDF (带样式)
- JSON (结构化数据)
- HTML (带图片)
- Excel (数据分析，多 sheet 工作簿)

作者: Claude Code
版本: 2.4.0
"""

import sys
import os
import json
import re
import argparse
import html
import logging
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime

# 配置日志
logger = logging.getLogger('wechat-exporter')


def _estimate_reading_time(text: str, wpm: int = 300) -> int:
    """
    估算阅读时间（分钟）

    Args:
        text: 文本内容
        wpm: 每分钟阅读字数（中文默认 300 字/分钟）

    Returns:
        int: 预计阅读分钟数（至少 1 分钟）
    """
    if not text:
        return 1
    # 中文字符计数
    chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', text))
    # 英文单词计数（近似）
    english_words = len(re.findall(r'[a-zA-Z]+', text))
    # 总字数 = 中文字符 + 英文单词
    total_words = chinese_chars + english_words
    minutes = max(1, round(total_words / wpm))
    return minutes


def _calculate_content_hash(data: Dict[str, Any]) -> str:
    """
    计算内容哈希，用于变更检测

    基于标题、作者、正文内容计算 SHA256 哈希
    """
    import hashlib

    # 提取关键字段
    title = data.get('title', '')
    author = data.get('author', '')
    content = data.get('content', '') or data.get('text', '')
    publish_time = data.get('publishTime', '')

    # 组合关键内容
    key_content = f"{title}|{author}|{publish_time}|{content[:5000]}"

    # 计算 SHA256
    return hashlib.sha256(key_content.encode('utf-8')).hexdigest()[:32]


def _calculate_wci(engagement: Dict[str, Any]) -> Optional[float]:
    """
    计算 WCI (WeChat Communication Index) 微信传播指数

    WCI 是衡量微信公众号文章传播力的指标，基于以下数据加权计算：
    - 阅读量 (R): 权重 50%
    - 点赞数 (L): 权重 20%
    - 在看数 (W): 权重 20%
    - 评论数 (C): 权重 10%

    公式参考: WCI = 0.5*ln(R+1) + 0.2*ln(L+1) + 0.2*ln(W+1) + 0.1*ln(C+1)
    然后归一化到 0-1000 范围

    Args:
        engagement: 互动数据字典，包含 readCount, likeCount, watchCount, commentCount

    Returns:
        Optional[float]: WCI 指数 (0-1000)，如果数据不足则返回 None
    """
    import math

    if not engagement:
        return None

    # 提取数值（处理字符串和 None）
    def parse_count(value) -> int:
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        # 处理 "1.2万" 格式
        if isinstance(value, str):
            value = value.strip()
            if '万' in value:
                try:
                    return int(float(value.replace('万', '')) * 10000)
                except ValueError:
                    return 0
            try:
                return int(value)
            except ValueError:
                return 0
        return 0

    read_count = parse_count(engagement.get('readCount'))
    like_count = parse_count(engagement.get('likeCount'))
    watch_count = parse_count(engagement.get('watchCount'))
    comment_count = parse_count(engagement.get('commentCount'))

    # 至少需要阅读量才能计算
    if read_count == 0:
        return None

    # 计算 WCI（对数加权）
    wci_raw = (
        0.5 * math.log(read_count + 1) +
        0.2 * math.log(like_count + 1) +
        0.2 * math.log(watch_count + 1) +
        0.1 * math.log(comment_count + 1)
    )

    # 归一化到 0-1000 范围（基于典型值调整）
    # ln(100000) ≈ 11.5，我们希望 10万阅读的文章 WCI 约 800-900
    wci_normalized = min(1000, round(wci_raw * 70))

    return wci_normalized


def _generate_summary(content: str, max_length: int = 200) -> str:
    """
    生成文章内容摘要

    使用简单的提取式摘要：
    1. 优先提取第一段有意义的文字
    2. 如果没有好的第一段，提取前 max_length 个字符
    3. 避免提取标题、作者信息、广告等噪音

    Args:
        content: 文章正文内容
        max_length: 摘要最大长度

    Returns:
        str: 文章摘要
    """
    if not content:
        return ""

    # 清理内容
    content = content.strip()

    # 按段落分割
    paragraphs = [p.strip() for p in content.split('\n') if p.strip()]

    # 噪音标记（避免作为摘要的段落开头）
    noise_prefixes = [
        '作者', '编辑', '来源', '转自', '公众号', '点击', '关注',
        '扫码', '二维码', '推荐阅读', '相关阅读', '往期回顾',
        '声明', '版权', '转载', '如有侵权', '商务合作',
        '👆', '↑', '↓', '【', '](http', '!['
    ]

    # 寻找合适的段落作为摘要
    for para in paragraphs[:5]:  # 只在前5段中找
        # 跳过太短的段落
        if len(para) < 30:
            continue

        # 跳过包含噪音标记的段落
        if any(para.startswith(prefix) for prefix in noise_prefixes):
            continue

        # 跳过主要是特殊字符的段落
        chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', para))
        if chinese_chars < 10:  # 中文字符太少
            continue

        # 找到合适的段落，截取前 max_length 字符
        if len(para) <= max_length:
            return para
        else:
            # 在 max_length 附近找句号、问号或感叹号截断
            truncate_pos = max_length
            for punct in ['。', '？', '！', '.', '?', '!']:
                pos = para.rfind(punct, max_length // 2, max_length)
                if pos > 0:
                    truncate_pos = pos + 1
                    break
            return para[:truncate_pos]

    # 如果没有找到合适的段落，直接截取开头
    if len(content) <= max_length:
        return content
    return content[:max_length] + "..."


def _clean_text(text: str) -> str:
    """
    清理文本内容

    吸取 wechat-article-browseruse 精华：
    - 处理 \xa0 非断空格（微信文章常见）
    - 规范化空白字符
    """
    if not text:
        return ""
    # 替换非断空格为普通空格
    text = text.replace("\xa0", " ")
    # 规范化空白字符
    text = re.sub(r"\s+", " ", text)
    return text.strip()


class Exporter:
    """文章导出器"""

    def __init__(self, output_dir: str = "."):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_markdown(
        self,
        data: Dict[str, Any],
        include_images: bool = True,
        include_meta: bool = True,
        converter: str = 'default'
    ) -> str:
        """
        导出为 Markdown

        Args:
            data: 文章数据字典
            include_images: 是否包含图片
            include_meta: 是否包含元数据头部
            converter: HTML 转换器选择
                - 'default': 使用原始内容
                - 'markdownify': 使用 markdownify 库
                - 'html2text': 使用 html2text 库（更轻量）

        Returns:
            str: Markdown 内容
        """
        # 根据 converter 选择转换方式
        if data.get('html'):
            if converter == 'markdownify':
                content = self._html_to_markdown_with_markdownify(data['html'])
            elif converter == 'html2text':
                content = self._html_to_markdown_with_html2text(data['html'])
            else:
                content = data.get('content', '') or data.get('text', '')
        else:
            content = data.get('content', '') or data.get('text', '')

        lines = []

        # 清理文本字段 - 吸取精华：处理 \xa0 非断空格
        title = _clean_text(data.get('title', '无标题'))
        author = _clean_text(data.get('author', '未知'))
        publish_time = _clean_text(data.get('publishTime', ''))
        source_url = data.get('source_url', '')
        description = _clean_text(data.get('description', ''))

        # 计算阅读时间
        content_for_reading = data.get('content', '') or data.get('text', '')
        reading_time = _estimate_reading_time(content_for_reading)

        # YAML Front Matter
        if include_meta:
            lines.append("---")
            lines.append(f"title: {title}")
            lines.append(f"author: {author}")
            if publish_time:
                lines.append(f"publish_time: {publish_time}")
            if source_url:
                lines.append(f"source_url: {source_url}")
            lines.append(f"exported_at: {datetime.now().isoformat()}")
            lines.append(f"reading_time: {reading_time} 分钟")
            if description:
                lines.append(f"description: {description}")
            lines.append("---")
            lines.append("")

        # 标题
        lines.append(f"# {title}")
        lines.append("")

        # 计算 WCI
        engagement = data.get('engagement', {})
        wci_score = _calculate_wci(engagement)

        # 元数据表格
        lines.append("**作者**: {}".format(author))
        if publish_time:
            lines.append("**发布时间**: {}".format(publish_time))
        # 新增：阅读时间
        lines.append("**阅读时间**: 约 {} 分钟".format(reading_time))
        # 新增：WCI 传播指数
        if wci_score:
            wci_level = "🔥爆款" if wci_score >= 800 else ("📈热门" if wci_score >= 500 else "📊普通")
            lines.append("**WCI传播指数**: {} ({})".format(wci_score, wci_level))
        # 新增：互动数据
        if engagement:
            if engagement.get('readCount'):
                lines.append("**阅读量**: {}".format(engagement['readCount']))
            if engagement.get('likeCount'):
                lines.append("**点赞数**: {}".format(engagement['likeCount']))
            if engagement.get('watchCount'):
                lines.append("**在看数**: {}".format(engagement['watchCount']))
        if data.get('source_url'):
            lines.append("**原文链接**: {}".format(data.get('source_url')))
        lines.append("")
        lines.append("---")
        lines.append("")

        # 处理内容中的图片
        if include_images and data.get('images'):
            content = self._insert_images_to_content(content, data['images'])

        lines.append(content)
        lines.append("")
        lines.append("---")
        lines.append("")

        # 图片列表
        if include_images and data.get('images'):
            lines.append("## 图片列表")
            lines.append("")
            for i, img in enumerate(data['images'], 1):
                src = img.get('src') or img.get('url', '')
                alt = img.get('alt', '')
                lines.append(f"{i}. ![{alt}]({src})")
            lines.append("")

        # 视频列表 - 新增功能
        if data.get('videos'):
            lines.append("## 视频列表")
            lines.append("")
            for i, video in enumerate(data['videos'], 1):
                src = video.get('src', '')
                poster = video.get('poster', '')
                title = video.get('title', '')
                duration = video.get('duration', '')
                info = f"{title} ({duration})" if title and duration else (title or '视频')
                lines.append(f"{i}. [{info}]({src or poster})")
            lines.append("")

        # 页脚
        lines.append(f"*本文档由 wechat-article-scraper 于 {datetime.now().strftime('%Y-%m-%d %H:%M')} 生成*")

        return '\n'.join(lines)

    def _html_to_markdown_with_markdownify(self, html: str) -> str:
        """
        使用 markdownify 将 HTML 转换为 Markdown

        竞品推荐此库，但以下问题需要验证:
        1. 中文排版支持是否更好
        2. 图片处理是否符合微信文章特征
        3. 性能 overhead 是否可接受
        """
        try:
            import markdownify

            # 微信特定的转换配置
            md = markdownify.markdownify(
                html,
                heading_style="ATX",  # # 样式的标题
                bullets="-",          # 统一使用 - 作为列表标记
                strip=['script', 'style', 'nav', 'header', 'footer'],
                convert=['b', 'i', 'strong', 'em', 'a', 'img', 'p', 'br', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'ul', 'ol', 'li', 'blockquote', 'code', 'pre', 'table', 'thead', 'tbody', 'tr', 'th', 'td']
            )

            # 后处理: 清理多余空行
            md = re.sub(r'\n{3,}', '\n\n', md)

            return md

        except ImportError:
            logger.warning("markdownify 未安装，使用默认转换")
            # 简单 fallback
            return html

    def _html_to_markdown_with_html2text(self, html: str) -> str:
        """
        使用 html2text 将 HTML 转换为 Markdown

        竞品 fetch-wx-article 使用此库，特点:
        1. 纯 Python 实现，更轻量
        2. 无额外依赖
        3. 转换速度快
        """
        try:
            import html2text

            h = html2text.HTML2Text()
            h.ignore_links = False  # 保留链接
            h.ignore_images = False  # 保留图片
            h.body_width = 0  # 不限制行宽

            return h.handle(html)

        except ImportError:
            logger.warning("html2text 未安装，使用默认转换")
            return html

    def _insert_images_to_content(self, content: str, images: list) -> str:
        """将图片插入到内容合适位置"""
        # 简单策略：在内容末尾添加图片
        # 更复杂的策略需要解析 HTML 结构
        return content

    def export_meta_sidecar(self, data: Dict[str, Any]) -> str:
        """
        导出元数据 sidecar 文件 (.meta.json)

        包含丰富的文章元数据，便于后续分析和索引：
        - 文章基本信息（标题、作者、发布时间）
        - 内容统计（字数、阅读时间）
        - 媒体资源统计（图片数、视频数）
        - 内容指纹（用于变更检测）
        - 抓取信息（策略、时间、状态）
        - 互动数据（阅读量、点赞数等）

        Args:
            data: 文章数据字典

        Returns:
            str: JSON 格式的元数据内容
        """
        content = data.get('content', '') or data.get('text', '')

        # 统计中文字符
        chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', content))
        english_words = len(re.findall(r'[a-zA-Z]+', content))

        # 计算阅读时间
        reading_time = _estimate_reading_time(content)

        # 计算内容哈希
        content_hash = _calculate_content_hash(data)

        # 图片统计
        images = data.get('images', [])
        image_stats = {
            'total': len(images),
            'with_alt': sum(1 for img in images if img.get('alt')),
            'local_paths': [img.get('local_path') for img in images if img.get('local_path')]
        }

        # 视频统计
        videos = data.get('videos', [])
        video_stats = {
            'total': len(videos),
            'durations': [v.get('duration') for v in videos if v.get('duration')]
        }

        # 互动数据
        engagement = data.get('engagement', {})

        # 计算 WCI (WeChat Communication Index)
        wci_score = _calculate_wci(engagement)

        # 生成摘要
        summary = _generate_summary(content)

        # 计算互动率（如果可能）
        engagement_rate = None
        if engagement and 'readCount' in engagement:
            read_count = 0
            rc = engagement.get('readCount')
            if isinstance(rc, (int, float)):
                read_count = int(rc)
            elif isinstance(rc, str) and '万' in rc:
                try:
                    read_count = int(float(rc.replace('万', '')) * 10000)
                except ValueError:
                    pass

            total_interactions = sum([
                int(engagement.get('likeCount', 0) or 0),
                int(engagement.get('watchCount', 0) or 0),
                int(engagement.get('commentCount', 0) or 0),
            ])

            if read_count > 0:
                engagement_rate = round((total_interactions / read_count) * 100, 2)

        meta = {
            'article': {
                'title': data.get('title', ''),
                'author': data.get('author', ''),
                'publish_time': data.get('publishTime', ''),
                'source_url': data.get('source_url', ''),
                'description': data.get('description', ''),
                'summary': summary,
            },
            'content': {
                'hash': content_hash,
                'chinese_chars': chinese_chars,
                'english_words': english_words,
                'total_words': chinese_chars + english_words,
                'reading_time_minutes': reading_time,
                'paragraphs': len(data.get('paragraphs', [])),
            },
            'media': {
                'images': image_stats,
                'videos': video_stats,
            },
            'engagement': {
                'data': engagement if engagement else None,
                'wci_score': wci_score,  # WeChat Communication Index
                'engagement_rate': engagement_rate,  # 互动率 %
            },
            'extraction': {
                'strategy': data.get('strategy', 'unknown'),
                'status': data.get('content_status', 'unknown'),
                'quality_score': data.get('quality_score'),
                'extracted_at': datetime.now().isoformat(),
                'version': '3.8.0',
            },
            'export': {
                'format': 'meta_sidecar',
                'exported_at': datetime.now().isoformat(),
            }
        }

        return json.dumps(meta, ensure_ascii=False, indent=2)

    def export_pdf(self, data: Dict[str, Any], output_file: str) -> str:
        """
        导出为 PDF

        需要安装 playwright 或 weasyprint
        """
        try:
            # 先导出 HTML
            html_content = self.export_html(data)

            # 使用 playwright 转换为 PDF
            from playwright.sync_api import sync_playwright

            with sync_playwright() as p:
                browser = p.chromium.launch()
                page = browser.new_page()

                # 加载 HTML
                page.set_content(html_content)

                # 等待图片加载
                page.wait_for_timeout(2000)

                # 生成 PDF
                page.pdf(
                    path=output_file,
                    format='A4',
                    margin={'top': '1cm', 'right': '1cm', 'bottom': '1cm', 'left': '1cm'},
                    print_background=True
                )

                browser.close()

            return output_file

        except ImportError:
            logger.error("导出 PDF 需要安装 playwright")
            logger.error("运行: pip install playwright && playwright install chromium")
            raise

    def export_json(self, data: Dict[str, Any]) -> str:
        """导出为 JSON"""
        # 添加导出元数据
        export_data = {
            **data,
            '_export_meta': {
                'version': '2.1.0',
                'exported_at': datetime.now().isoformat(),
                'exporter': 'wechat-article-scraper'
            }
        }
        return json.dumps(export_data, ensure_ascii=False, indent=2)

    def export_html(self, data: Dict[str, Any]) -> str:
        """导出为 HTML"""
        # 使用 html.escape() 对所有用户输入进行转义，防止 XSS
        title = html.escape(data.get('title', '无标题'))
        author = html.escape(data.get('author', '未知'))
        content = data.get('html', '') or data.get('content', '') or data.get('text', '')
        source_url = html.escape(data.get('source_url', '#'))
        publish_time = html.escape(data.get('publishTime', ''))

        # 如果没有 HTML，将文本转换为简单 HTML
        if not content.startswith('<'):
            # 先转义文本内容，再替换换行符
            content = html.escape(content)
            content = f"<p>{content.replace(chr(10), '</p><p>')}</p>"
        else:
            # 如果已有 HTML，清理潜在的 XSS
            content = self._sanitize_html(content)

        # 时间标签 HTML（已转义）
        time_span = f'<span>&#128197; {publish_time}</span>' if publish_time else ''

        # 视频部分 HTML（新增）
        video_section = ''
        if data.get('videos'):
            video_items = []
            for video in data['videos']:
                src = html.escape(video.get('src', ''))
                poster = html.escape(video.get('poster', ''))
                title = html.escape(video.get('title', '视频'))
                duration = html.escape(video.get('duration', ''))

                # 使用 video 标签或链接
                if src:
                    video_html = f'<video controls preload="metadata" poster="{poster}" style="max-width:100%;margin:20px 0;"><source src="{src}"></video>'
                elif poster:
                    video_html = f'<div style="margin:20px 0;"><img src="{poster}" style="max-width:100%;" alt="{title}"><p>{title} (视频)</p></div>'
                else:
                    continue

                info = f"{title} ({duration})" if duration else title
                video_items.append(f'<div style="margin:20px 0;"><p><strong>{info}</strong></p>{video_html}</div>')

            if video_items:
                video_section = f'<div class="videos" style="margin-top:40px;padding-top:20px;border-top:1px solid #eee;"><h2>视频列表</h2>{"".join(video_items)}</div>'

        html_template = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            line-height: 1.8;
            max-width: 800px;
            margin: 0 auto;
            padding: 40px 20px;
            color: #333;
            background: #f5f5f5;
        }}
        .container {{
            background: white;
            padding: 40px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        h1 {{
            font-size: 28px;
            margin-bottom: 10px;
            color: #1a1a1a;
        }}
        .meta {{
            color: #666;
            font-size: 14px;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 1px solid #eee;
        }}
        .meta span {{
            margin-right: 20px;
        }}
        .content {{
            font-size: 16px;
        }}
        .content p {{
            margin: 1em 0;
        }}
        .content img {{
            max-width: 100%;
            height: auto;
            display: block;
            margin: 20px auto;
            border-radius: 4px;
        }}
        .footer {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #eee;
            color: #999;
            font-size: 12px;
            text-align: center;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        <div class="meta">
            <span>&#128100; {author}</span>
            {time_span}
        </div>
        <div class="content">
            {content}
        </div>
        {video_section}
        <div class="footer">
            原文链接: <a href="{source_url}" target="_blank">{source_url}</a><br>
            导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
        </div>
    </div>
</body>
</html>"""

        return html_template

    def _sanitize_html(self, html_content: str) -> str:
        """清理 HTML 内容中的潜在 XSS（强化版本）"""
        import re

        # 限制输入长度，防止 ReDoS 攻击
        max_length = 10 * 1024 * 1024  # 10MB 上限
        if len(html_content) > max_length:
            html_content = html_content[:max_length]

        # 移除危险标签（非贪婪匹配）
        dangerous_tags = ['script', 'iframe', 'object', 'embed', 'form', 'input', 'textarea', 'select']
        for tag in dangerous_tags:
            # 移除开始标签（非贪婪匹配）
            html_content = re.sub(
                rf'<{tag}\b[^>]*?>',
                f'&lt;{tag}&gt;',
                html_content,
                flags=re.IGNORECASE
            )
            # 移除结束标签
            html_content = re.sub(
                rf'</{tag}\s*>',
                f'&lt;/{tag}&gt;',
                html_content,
                flags=re.IGNORECASE
            )

        # 移除事件处理器（非贪婪匹配）
        html_content = re.sub(
            r'\son\w+\s*=\s*["\'][^"\']*?["\']',
            '',
            html_content,
            flags=re.IGNORECASE
        )

        # 移除 javascript: 和 data: 伪协议
        html_content = re.sub(
            r'(href|src|action)\s*=\s*["\']\s*(?:javascript|data):[^"\']*?["\']',
            r'\1="#"',
            html_content,
            flags=re.IGNORECASE
        )

        # 移除 CSS expression（IE 特有的 XSS 攻击向量）
        html_content = re.sub(
            r'expression\s*\(',
            'expression_removed(',
            html_content,
            flags=re.IGNORECASE
        )

        # 移除 style 标签中的 @import 和 behavior
        html_content = re.sub(
            r'@import\s+["\']',
            '@import_removed "',
            html_content,
            flags=re.IGNORECASE
        )
        html_content = re.sub(
            r'behavior\s*:',
            'behavior_removed:',
            html_content,
            flags=re.IGNORECASE
        )

        return html_content

    def export_excel(self, articles: list, output_file: str) -> str:
        """
        导出为 Excel (多 sheet 工作簿)

        吸取竞品 wcplusPro 精华：数据分析师需要结构化表格数据

        Sheet 结构:
        1. 文章列表 - 批量文章的元数据概览
        2. 互动数据 - 阅读量、点赞数、在看数、WCI 指数
        3. 分类统计 - 自动分类结果汇总
        4. 媒体资源 - 图片、视频统计
        5. 详细数据 - 单篇文章的完整信息（如果只有一篇）

        Args:
            articles: 文章数据列表，每个元素是文章字典
            output_file: 输出文件路径

        Returns:
            str: 输出文件路径

        Raises:
            ImportError: 如果 openpyxl 未安装
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("导出 Excel 需要安装 openpyxl")
            logger.error("运行: pip install openpyxl")
            raise

        # 确保 articles 是列表
        if isinstance(articles, dict):
            articles = [articles]

        wb = openpyxl.Workbook()

        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def style_header(ws, row_num=1):
            """为表头行应用样式"""
            for cell in ws[row_num]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

        def auto_column_width(ws):
            """自动调整列宽"""
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = min(cell_length, 50)  # 最大50字符
                    except:
                        pass
                adjusted_width = max(10, min(max_length + 2, 60))
                ws.column_dimensions[column_letter].width = adjusted_width

        # ==================== Sheet 1: 文章列表 ====================
        ws1 = wb.active
        ws1.title = "文章列表"

        headers1 = ["序号", "标题", "作者", "发布时间", "字数", "阅读时间(分)", "图片数", "视频数", "原文链接"]
        ws1.append(headers1)

        for idx, article in enumerate(articles, 1):
            content = article.get('content', '') or article.get('text', '')
            chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', content))
            english_words = len(re.findall(r'[a-zA-Z]+', content))
            total_words = chinese_chars + english_words
            reading_time = _estimate_reading_time(content)

            row = [
                idx,
                article.get('title', ''),
                article.get('author', ''),
                article.get('publishTime', ''),
                total_words,
                reading_time,
                len(article.get('images', [])),
                len(article.get('videos', [])),
                article.get('source_url', '')
            ]
            ws1.append(row)

        style_header(ws1)
        auto_column_width(ws1)

        # ==================== Sheet 2: 互动数据 ====================
        ws2 = wb.create_sheet("互动数据")

        headers2 = ["序号", "标题", "作者", "阅读量", "点赞数", "在看数", "评论数", "WCI指数", "互动率%", "传播等级"]
        ws2.append(headers2)

        for idx, article in enumerate(articles, 1):
            engagement = article.get('engagement', {})

            # 解析互动数据
            def parse_count(val):
                if val is None:
                    return 0
                if isinstance(val, (int, float)):
                    return int(val)
                if isinstance(val, str):
                    if '万' in val:
                        try:
                            return int(float(val.replace('万', '')) * 10000)
                        except:
                            return 0
                    try:
                        return int(val)
                    except:
                        return 0
                return 0

            read_count = parse_count(engagement.get('readCount'))
            like_count = parse_count(engagement.get('likeCount'))
            watch_count = parse_count(engagement.get('watchCount'))
            comment_count = parse_count(engagement.get('commentCount'))

            # 计算 WCI
            wci = _calculate_wci(engagement)

            # 计算互动率
            engagement_rate = 0
            if read_count > 0:
                engagement_rate = round(((like_count + watch_count + comment_count) / read_count) * 100, 2)

            # 传播等级
            if wci:
                level = "🔥爆款" if wci >= 800 else ("📈热门" if wci >= 500 else ("📊良好" if wci >= 300 else "📉普通"))
            else:
                level = "未知"

            row = [
                idx,
                article.get('title', ''),
                article.get('author', ''),
                read_count if read_count > 0 else "-",
                like_count if like_count > 0 else "-",
                watch_count if watch_count > 0 else "-",
                comment_count if comment_count > 0 else "-",
                wci if wci else "-",
                f"{engagement_rate}%" if read_count > 0 else "-",
                level
            ]
            ws2.append(row)

        style_header(ws2)
        auto_column_width(ws2)

        # ==================== Sheet 3: 分类统计 ====================
        ws3 = wb.create_sheet("分类统计")

        # 统计分类分布
        categories = {}
        for article in articles:
            cat = article.get('category', '未分类')
            if isinstance(cat, dict):
                cat = cat.get('primary_category', '未分类')
            if cat not in categories:
                categories[cat] = {'count': 0, 'articles': []}
            categories[cat]['count'] += 1
            categories[cat]['articles'].append(article.get('title', ''))

        headers3 = ["分类", "文章数量", "占比%", "文章列表"]
        ws3.append(headers3)

        total = len(articles)
        for cat, data in sorted(categories.items(), key=lambda x: x[1]['count'], reverse=True):
            percentage = round(data['count'] / total * 100, 2) if total > 0 else 0
            article_list = "; ".join(data['articles'][:5])
            if len(data['articles']) > 5:
                article_list += f"; ...等{data['count'] - 5}篇"

            ws3.append([cat, data['count'], percentage, article_list])

        style_header(ws3)
        auto_column_width(ws3)

        # ==================== Sheet 4: 媒体资源 ====================
        ws4 = wb.create_sheet("媒体资源")

        headers4 = ["序号", "标题", "图片数量", "视频数量", "图片URL列表", "视频信息"]
        ws4.append(headers4)

        for idx, article in enumerate(articles, 1):
            images = article.get('images', [])
            videos = article.get('videos', [])

            # 图片 URL 列表（前5个）
            image_urls = [img.get('src', '') for img in images[:5]]
            image_str = "; ".join(image_urls)
            if len(images) > 5:
                image_str += f"; ...等{len(images) - 5}张"

            # 视频信息
            video_info = []
            for v in videos[:3]:
                title = v.get('title', '')
                duration = v.get('duration', '')
                info = f"{title}({duration})" if title and duration else (title or '未命名')
                video_info.append(info)
            video_str = "; ".join(video_info)
            if len(videos) > 3:
                video_str += f"; ...等{len(videos) - 3}个"

            ws4.append([
                idx,
                article.get('title', ''),
                len(images),
                len(videos),
                image_str,
                video_str
            ])

        style_header(ws4)
        auto_column_width(ws4)

        # ==================== Sheet 5: 详细数据（仅单篇文章） ====================
        if len(articles) == 1:
            ws5 = wb.create_sheet("详细数据")
            article = articles[0]
            content = article.get('content', '') or article.get('text', '')

            # 基本信息
            ws5.append(["字段", "值"])
            ws5.append(["标题", article.get('title', '')])
            ws5.append(["作者", article.get('author', '')])
            ws5.append(["发布时间", article.get('publishTime', '')])
            ws5.append(["原文链接", article.get('source_url', '')])
            ws5.append([])

            # 内容统计
            chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', content))
            english_words = len(re.findall(r'[a-zA-Z]+', content))
            ws5.append(["中文字符数", chinese_chars])
            ws5.append(["英文单词数", english_words])
            ws5.append(["总字数", chinese_chars + english_words])
            ws5.append(["预计阅读时间(分钟)", _estimate_reading_time(content)])
            ws5.append(["段落数", len(article.get('paragraphs', []))])
            ws5.append([])

            # 互动数据
            engagement = article.get('engagement', {})
            ws5.append(["阅读量", engagement.get('readCount', '-')])
            ws5.append(["点赞数", engagement.get('likeCount', '-')])
            ws5.append(["在看数", engagement.get('watchCount', '-')])
            ws5.append(["评论数", engagement.get('commentCount', '-')])
            wci = _calculate_wci(engagement)
            ws5.append(["WCI传播指数", wci if wci else '-'])
            ws5.append([])

            # 正文内容（限制长度）
            ws5.append(["正文内容", content[:3000] + "..." if len(content) > 3000 else content])

            # 样式
            ws5['A1'].font = Font(bold=True)
            ws5['A1'].fill = header_fill
            ws5['B1'].font = Font(bold=True)
            ws5['B1'].fill = header_fill
            ws5.column_dimensions['A'].width = 20
            ws5.column_dimensions['B'].width = 80

        # 保存文件
        wb.save(output_file)
        logger.info(f"Excel 导出成功: {output_file}")
        return output_file

    def save(
        self,
        data: Dict[str, Any],
        format: str,
        filename: Optional[str] = None,
        converter: str = 'default',
        include_sidecar: bool = False
    ) -> str:
        """
        保存文章到文件

        Args:
            data: 文章数据
            format: 格式 (markdown, pdf, json, html)
            filename: 文件名（不含扩展名）
            converter: Markdown 转换器选择（'default'/'markdownify'/'html2text'）
            include_sidecar: 是否同时生成元数据 sidecar 文件 (.meta.json)

        Returns:
            str: 保存的文件路径
        """
        # 生成文件名
        if not filename:
            title = data.get('title', 'untitled')
            # 清理文件名中的非法字符
            filename = re.sub(r'[<>"/\\|?*]', '', title)[:50]

        # 根据格式选择导出方法
        format_methods = {
            'markdown': (self.export_markdown, 'md'),
            'md': (self.export_markdown, 'md'),
            'json': (self.export_json, 'json'),
            'html': (self.export_html, 'html'),
            'pdf': (self.export_pdf, 'pdf'),
            'excel': (self.export_excel, 'xlsx'),
            'xlsx': (self.export_excel, 'xlsx'),
        }

        if format not in format_methods:
            raise ValueError(f"不支持的格式: {format}。支持: {list(format_methods.keys())}")

        method, ext = format_methods[format]

        # Excel 和 PDF 需要特殊处理
        if format in ('pdf', 'excel', 'xlsx'):
            output_path = self.output_dir / f"{filename}.{ext}"
            method(data if format == 'pdf' else ([data] if isinstance(data, dict) else data), str(output_path))
        elif format in ('markdown', 'md'):
            output_path = self.output_dir / f"{filename}.{ext}"
            content = method(data, converter=converter)
            output_path.write_text(content, encoding='utf-8')
        else:
            output_path = self.output_dir / f"{filename}.{ext}"
            content = method(data)
            output_path.write_text(content, encoding='utf-8')

        # 生成元数据 sidecar 文件
        if include_sidecar:
            sidecar_path = self.output_dir / f"{filename}.meta.json"
            sidecar_content = self.export_meta_sidecar(data)
            sidecar_path.write_text(sidecar_content, encoding='utf-8')
            logger.info(f"元数据 sidecar 已保存: {sidecar_path}")

        return str(output_path)


def main():
    parser = argparse.ArgumentParser(
        description='将微信文章数据导出为多种格式'
    )
    parser.add_argument(
        'input',
        help='输入文件 (JSON 格式)'
    )
    parser.add_argument(
        '-f', '--format',
        choices=['markdown', 'md', 'json', 'html', 'pdf', 'excel', 'xlsx'],
        default='markdown',
        help='输出格式 (默认: markdown)'
    )
    parser.add_argument(
        '-o', '--output',
        help='输出文件路径'
    )
    parser.add_argument(
        '-d', '--dir',
        default='.',
        help='输出目录'
    )
    parser.add_argument(
        '--sidecar',
        action='store_true',
        help='同时生成元数据 sidecar 文件 (.meta.json)'
    )
    parser.add_argument(
        '--meta-only',
        action='store_true',
        help='仅生成元数据 sidecar 文件'
    )

    args = parser.parse_args()

    # 读取输入
    input_path = Path(args.input)
    if not input_path.exists():
        logger.error(f"文件不存在 {args.input}")
        sys.exit(1)

    data = json.loads(input_path.read_text(encoding='utf-8'))

    # 导出
    exporter = Exporter(output_dir=args.dir)

    try:
        if args.meta_only:
            # 仅生成元数据 sidecar
            if not args.output:
                title = data.get('title', 'untitled')
                args.output = re.sub(r'[<>"/\\|?*]', '', title)[:50]
            sidecar_path = exporter.output_dir / f"{args.output}.meta.json"
            sidecar_content = exporter.export_meta_sidecar(data)
            sidecar_path.write_text(sidecar_content, encoding='utf-8')
            print(f"元数据 sidecar 导出成功: {sidecar_path}")
        else:
            output_path = exporter.save(
                data,
                format=args.format,
                filename=args.output,
                include_sidecar=args.sidecar
            )
            print(f"导出成功: {output_path}")
            if args.sidecar:
                sidecar_path = exporter.output_dir / f"{Path(output_path).stem}.meta.json"
                print(f"元数据 sidecar: {sidecar_path}")
    except Exception as e:
        print(f"导出失败: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
