#!/usr/bin/env python3
"""
SQLite 持久化存储模块

功能：
- 文章数据持久化存储 (SQLite)
- 内容变更检测 (基于 SHA256 哈希)
- 增量更新支持 (避免重复抓取)
- SQL 查询接口 (历史数据分析)
- 全文搜索支持 (标题和内容)

吸取竞品 wcplusPro 精华：
- 持久化存储是数据分析的基础
- 增量更新节省时间和带宽
- 内容指纹检测变更

作者: Claude Code
版本: 1.0.0
"""

import re
import json
import sqlite3
import hashlib
import logging
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime, timedelta
from dataclasses import dataclass, asdict
from contextlib import contextmanager

logger = logging.getLogger('wechat-storage')


@dataclass
class ArticleRecord:
    """文章记录数据类"""
    id: Optional[int] = None
    url: str = ""
    title: str = ""
    author: str = ""
    publish_time: str = ""
    content: str = ""
    content_hash: str = ""  # 内容指纹，用于变更检测
    html_content: str = ""
    source_url: str = ""

    # 媒体资源 (JSON 存储)
    images_json: str = "[]"
    videos_json: str = "[]"

    # 互动数据 (JSON 存储)
    engagement_json: str = "{}"
    wci_score: Optional[int] = None

    # 分类信息
    category: str = ""
    category_confidence: float = 0.0

    # 元数据
    strategy: str = ""
    quality_score: Optional[float] = None
    content_status: str = ""

    # 时间戳
    created_at: str = ""
    updated_at: str = ""
    last_synced_at: str = ""


class ArticleStorage:
    """
    文章 SQLite 存储管理器

    吸取竞品精华：
    - wcplusPro 使用 MongoDB，我们使用 SQLite 实现零配置
    - 内容哈希实现增量更新（只抓取变更内容）
    - 全文搜索支持快速查找
    """

    def __init__(self, db_path: str = "wechat_articles.db"):
        """
        初始化存储管理器

        Args:
            db_path: SQLite 数据库文件路径
        """
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_database()

    @contextmanager
    def _get_connection(self):
        """获取数据库连接的上下文管理器"""
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row  # 使结果可以通过列名访问
        try:
            yield conn
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()

    def _init_database(self):
        """初始化数据库表结构"""
        with self._get_connection() as conn:
            cursor = conn.cursor()

            # 主表：文章数据
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS articles (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    url TEXT UNIQUE NOT NULL,
                    title TEXT NOT NULL,
                    author TEXT,
                    publish_time TEXT,
                    content TEXT,
                    content_hash TEXT NOT NULL,
                    html_content TEXT,
                    source_url TEXT,
                    images_json TEXT DEFAULT '[]',
                    videos_json TEXT DEFAULT '[]',
                    engagement_json TEXT DEFAULT '{}',
                    wci_score INTEGER,
                    category TEXT,
                    category_confidence REAL DEFAULT 0,
                    strategy TEXT,
                    quality_score REAL,
                    content_status TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    last_synced_at TEXT NOT NULL
                )
            """)

            # 索引优化查询
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_articles_url ON articles(url)
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_articles_author ON articles(author)
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_articles_category ON articles(category)
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_articles_publish_time ON articles(publish_time)
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_articles_created_at ON articles(created_at)
            """)

            # 全文搜索虚拟表
            cursor.execute("""
                CREATE VIRTUAL TABLE IF NOT EXISTS articles_fts USING fts5(
                    title, content, author, category,
                    content='articles',
                    content_rowid='id'
                )
            """)

            # 触发器：自动同步全文搜索
            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS articles_ai AFTER INSERT ON articles BEGIN
                    INSERT INTO articles_fts(rowid, title, content, author, category)
                    VALUES (new.id, new.title, new.content, new.author, new.category);
                END
            """)

            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS articles_ad AFTER DELETE ON articles BEGIN
                    INSERT INTO articles_fts(articles_fts, rowid, title, content, author, category)
                    VALUES ('delete', old.id, old.title, old.content, old.author, old.category);
                END
            """)

            cursor.execute("""
                CREATE TRIGGER IF NOT EXISTS articles_au AFTER UPDATE ON articles BEGIN
                    INSERT INTO articles_fts(articles_fts, rowid, title, content, author, category)
                    VALUES ('delete', old.id, old.title, old.content, old.author, old.category);
                    INSERT INTO articles_fts(rowid, title, content, author, category)
                    VALUES (new.id, new.title, new.content, new.author, new.category);
                END
            """)

            # 同步历史表：记录抓取历史
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS sync_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    article_id INTEGER NOT NULL,
                    sync_type TEXT NOT NULL,
                    old_hash TEXT,
                    new_hash TEXT,
                    synced_at TEXT NOT NULL,
                    FOREIGN KEY (article_id) REFERENCES articles(id)
                )
            """)

            logger.info(f"数据库初始化完成: {self.db_path}")

    def _calculate_content_hash(self, data: Dict[str, Any]) -> str:
        """计算内容哈希，用于变更检测"""
        key_content = f"{data.get('title', '')}|{data.get('author', '')}|{data.get('publishTime', '')}|{data.get('content', '')[:5000]}"
        return hashlib.sha256(key_content.encode('utf-8')).hexdigest()

    def save_article(self, data: Dict[str, Any]) -> Tuple[int, str]:
        """
        保存或更新文章

        Args:
            data: 文章数据字典

        Returns:
            Tuple[int, str]: (文章ID, 操作类型: 'inserted', 'updated', 'unchanged')
        """
        url = data.get('source_url', '') or data.get('url', '')
        if not url:
            raise ValueError("文章 URL 不能为空")

        content_hash = self._calculate_content_hash(data)
        now = datetime.now().isoformat()

        with self._get_connection() as conn:
            cursor = conn.cursor()

            # 检查是否已存在
            cursor.execute("SELECT id, content_hash FROM articles WHERE url = ?", (url,))
            existing = cursor.fetchone()

            # 准备数据
            record = {
                'url': url,
                'title': data.get('title', ''),
                'author': data.get('author', ''),
                'publish_time': data.get('publishTime', ''),
                'content': data.get('content', '') or data.get('text', ''),
                'content_hash': content_hash,
                'html_content': data.get('html', ''),
                'source_url': url,
                'images_json': json.dumps(data.get('images', []), ensure_ascii=False),
                'videos_json': json.dumps(data.get('videos', []), ensure_ascii=False),
                'engagement_json': json.dumps(data.get('engagement', {}), ensure_ascii=False),
                'wci_score': self._calculate_wci(data.get('engagement', {})),
                'category': data.get('category', {}).get('primary_category', '') if isinstance(data.get('category'), dict) else data.get('category', ''),
                'category_confidence': data.get('category', {}).get('confidence', 0) if isinstance(data.get('category'), dict) else 0,
                'strategy': data.get('strategy', ''),
                'quality_score': data.get('quality_score'),
                'content_status': data.get('content_status', ''),
                'updated_at': now,
                'last_synced_at': now
            }

            if existing:
                existing_id, existing_hash = existing

                if existing_hash == content_hash:
                    # 内容未变更，只更新同步时间
                    cursor.execute("""
                        UPDATE articles SET last_synced_at = ? WHERE id = ?
                    """, (now, existing_id))

                    # 记录同步历史
                    cursor.execute("""
                        INSERT INTO sync_history (article_id, sync_type, old_hash, new_hash, synced_at)
                        VALUES (?, 'unchanged', ?, ?, ?)
                    """, (existing_id, existing_hash, content_hash, now))

                    return existing_id, 'unchanged'
                else:
                    # 内容已变更，执行更新
                    record['created_at'] = now  # 保留原创建时间会更好，但这里简化处理
                    record['id'] = existing_id

                    cursor.execute("""
                        UPDATE articles SET
                            title = ?, author = ?, publish_time = ?, content = ?,
                            content_hash = ?, html_content = ?, images_json = ?,
                            videos_json = ?, engagement_json = ?, wci_score = ?,
                            category = ?, category_confidence = ?, strategy = ?,
                            quality_score = ?, content_status = ?, updated_at = ?,
                            last_synced_at = ?
                        WHERE id = ?
                    """, (
                        record['title'], record['author'], record['publish_time'],
                        record['content'], record['content_hash'], record['html_content'],
                        record['images_json'], record['videos_json'], record['engagement_json'],
                        record['wci_score'], record['category'], record['category_confidence'],
                        record['strategy'], record['quality_score'], record['content_status'],
                        record['updated_at'], record['last_synced_at'], existing_id
                    ))

                    # 记录同步历史
                    cursor.execute("""
                        INSERT INTO sync_history (article_id, sync_type, old_hash, new_hash, synced_at)
                        VALUES (?, 'updated', ?, ?, ?)
                    """, (existing_id, existing_hash, content_hash, now))

                    return existing_id, 'updated'
            else:
                # 新文章，执行插入
                record['created_at'] = now

                cursor.execute("""
                    INSERT INTO articles (
                        url, title, author, publish_time, content, content_hash,
                        html_content, source_url, images_json, videos_json,
                        engagement_json, wci_score, category, category_confidence,
                        strategy, quality_score, content_status, created_at,
                        updated_at, last_synced_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    record['url'], record['title'], record['author'],
                    record['publish_time'], record['content'], record['content_hash'],
                    record['html_content'], record['source_url'], record['images_json'],
                    record['videos_json'], record['engagement_json'], record['wci_score'],
                    record['category'], record['category_confidence'], record['strategy'],
                    record['quality_score'], record['content_status'], record['created_at'],
                    record['updated_at'], record['last_synced_at']
                ))

                new_id = cursor.lastrowid

                # 记录同步历史
                cursor.execute("""
                    INSERT INTO sync_history (article_id, sync_type, old_hash, new_hash, synced_at)
                    VALUES (?, 'inserted', NULL, ?, ?)
                """, (new_id, content_hash, now))

                return new_id, 'inserted'

    def get_article(self, url: str) -> Optional[Dict[str, Any]]:
        """通过 URL 获取文章"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM articles WHERE url = ?", (url,))
            row = cursor.fetchone()

            if row:
                return self._row_to_dict(row)
            return None

    def get_article_by_id(self, article_id: int) -> Optional[Dict[str, Any]]:
        """通过 ID 获取文章"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM articles WHERE id = ?", (article_id,))
            row = cursor.fetchone()

            if row:
                return self._row_to_dict(row)
            return None

    def search_articles(self, keyword: str, limit: int = 20) -> List[Dict[str, Any]]:
        """
        全文搜索文章

        Args:
            keyword: 搜索关键词
            limit: 返回数量限制

        Returns:
            文章列表，按相关性排序
        """
        with self._get_connection() as conn:
            cursor = conn.cursor()

            # 使用 FTS5 全文搜索
            cursor.execute("""
                SELECT a.* FROM articles a
                JOIN articles_fts fts ON a.id = fts.rowid
                WHERE articles_fts MATCH ?
                ORDER BY rank
                LIMIT ?
            """, (keyword, limit))

            return [self._row_to_dict(row) for row in cursor.fetchall()]

    def list_articles(
        self,
        author: Optional[str] = None,
        category: Optional[str] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        limit: int = 100,
        offset: int = 0
    ) -> List[Dict[str, Any]]:
        """
        列出文章（支持筛选）

        Args:
            author: 按作者筛选
            category: 按分类筛选
            start_date: 开始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD)
            limit: 返回数量
            offset: 偏移量
        """
        with self._get_connection() as conn:
            cursor = conn.cursor()

            conditions = []
            params = []

            if author:
                conditions.append("author = ?")
                params.append(author)

            if category:
                conditions.append("category = ?")
                params.append(category)

            if start_date:
                conditions.append("publish_time >= ?")
                params.append(start_date)

            if end_date:
                conditions.append("publish_time <= ?")
                params.append(end_date)

            where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

            query = f"""
                SELECT * FROM articles
                {where_clause}
                ORDER BY publish_time DESC
                LIMIT ? OFFSET ?
            """
            params.extend([limit, offset])

            cursor.execute(query, params)
            return [self._row_to_dict(row) for row in cursor.fetchall()]

    def get_statistics(self) -> Dict[str, Any]:
        """获取数据库统计信息"""
        with self._get_connection() as conn:
            cursor = conn.cursor()

            # 总文章数
            cursor.execute("SELECT COUNT(*) FROM articles")
            total_count = cursor.fetchone()[0]

            # 作者分布
            cursor.execute("""
                SELECT author, COUNT(*) as count FROM articles
                WHERE author != ''
                GROUP BY author
                ORDER BY count DESC
                LIMIT 10
            """)
            author_stats = [{'author': row[0], 'count': row[1]} for row in cursor.fetchall()]

            # 分类分布
            cursor.execute("""
                SELECT category, COUNT(*) as count FROM articles
                WHERE category != ''
                GROUP BY category
                ORDER BY count DESC
            """)
            category_stats = [{'category': row[0], 'count': row[1]} for row in cursor.fetchall()]

            # 最近同步统计
            cursor.execute("""
                SELECT sync_type, COUNT(*) as count FROM sync_history
                WHERE synced_at >= datetime('now', '-7 days')
                GROUP BY sync_type
            """)
            sync_stats = {row[0]: row[1] for row in cursor.fetchall()}

            # WCI 分布
            cursor.execute("""
                SELECT
                    CASE
                        WHEN wci_score >= 800 THEN '爆款(800+)'
                        WHEN wci_score >= 500 THEN '热门(500-799)'
                        WHEN wci_score >= 300 THEN '良好(300-499)'
                        ELSE '普通(<300)'
                    END as level,
                    COUNT(*) as count
                FROM articles
                WHERE wci_score IS NOT NULL
                GROUP BY level
                ORDER BY wci_score DESC
            """)
            wci_stats = [{'level': row[0], 'count': row[1]} for row in cursor.fetchall()]

            return {
                'total_articles': total_count,
                'top_authors': author_stats,
                'category_distribution': category_stats,
                'recent_sync': sync_stats,
                'wci_distribution': wci_stats,
                'database_path': str(self.db_path),
                'generated_at': datetime.now().isoformat()
            }

    def get_sync_history(self, url: str, limit: int = 10) -> List[Dict[str, Any]]:
        """获取文章的同步历史"""
        with self._get_connection() as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT h.* FROM sync_history h
                JOIN articles a ON h.article_id = a.id
                WHERE a.url = ?
                ORDER BY h.synced_at DESC
                LIMIT ?
            """, (url, limit))

            columns = [description[0] for description in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def delete_article(self, url: str) -> bool:
        """删除文章"""
        with self._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM articles WHERE url = ?", (url,))
            return cursor.rowcount > 0

    def export_to_excel(self, output_path: str, **filters) -> str:
        """
        导出数据到 Excel

        吸取竞品精华：数据分析师需要结构化表格
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("导出 Excel 需要 openpyxl: pip install openpyxl")

        articles = self.list_articles(**filters, limit=10000)

        if not articles:
            raise ValueError("没有符合条件的数据")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "文章数据"

        # 表头
        headers = ["ID", "标题", "作者", "发布时间", "分类", "WCI", "字数", "图片数", "视频数", "抓取策略", "最后同步"]
        ws.append(headers)

        # 样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 数据
        for article in articles:
            content = article.get('content', '')
            chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', content))
            english_words = len(re.findall(r'[a-zA-Z]+', content))

            images = json.loads(article.get('images_json', '[]'))
            videos = json.loads(article.get('videos_json', '[]'))

            ws.append([
                article.get('id'),
                article.get('title', ''),
                article.get('author', ''),
                article.get('publish_time', ''),
                article.get('category', ''),
                article.get('wci_score', ''),
                chinese_chars + english_words,
                len(images),
                len(videos),
                article.get('strategy', ''),
                article.get('last_synced_at', '')[:10] if article.get('last_synced_at') else ''
            ])

        # 自动调整列宽
        from openpyxl.utils import get_column_letter
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = min(len(str(cell.value)), 50)
                except:
                    pass
            ws.column_dimensions[column_letter].width = max(10, min(max_length + 2, 60))

        wb.save(output_path)
        logger.info(f"数据已导出到 Excel: {output_path}")
        return output_path

    def _calculate_wci(self, engagement: Dict[str, Any]) -> Optional[int]:
        """计算 WCI 传播指数"""
        import math

        if not engagement:
            return None

        def parse_count(val):
            if val is None:
                return 0
            if isinstance(val, (int, float)):
                return int(val)
            if isinstance(val, str):
                if '万' in val:
                    try:
                        return int(float(val.replace('万', '')) * 10000)
                    except:
                        return 0
                try:
                    return int(val)
                except:
                    return 0
            return 0

        read_count = parse_count(engagement.get('readCount'))
        like_count = parse_count(engagement.get('likeCount'))
        watch_count = parse_count(engagement.get('watchCount'))
        comment_count = parse_count(engagement.get('commentCount'))

        if read_count == 0:
            return None

        wci_raw = (
            0.5 * math.log(read_count + 1) +
            0.2 * math.log(like_count + 1) +
            0.2 * math.log(watch_count + 1) +
            0.1 * math.log(comment_count + 1)
        )

        return min(1000, round(wci_raw * 70))

    def _row_to_dict(self, row: sqlite3.Row) -> Dict[str, Any]:
        """将数据库行转换为字典"""
        result = dict(row)

        # 解析 JSON 字段
        for field in ['images_json', 'videos_json', 'engagement_json']:
            if field in result and result[field]:
                try:
                    result[field.replace('_json', '')] = json.loads(result[field])
                except:
                    result[field.replace('_json', '')] = []
            else:
                result[field.replace('_json', '')] = []

        return result


def main():
    """CLI 入口"""
    import argparse

    parser = argparse.ArgumentParser(description='微信文章 SQLite 存储管理')
    parser.add_argument('--db', default='wechat_articles.db', help='数据库文件路径')

    subparsers = parser.add_subparsers(dest='command', help='可用命令')

    # save 命令
    save_parser = subparsers.add_parser('save', help='保存文章')
    save_parser.add_argument('file', help='文章 JSON 文件')

    # get 命令
    get_parser = subparsers.add_parser('get', help='获取文章')
    get_parser.add_argument('url', help='文章 URL')

    # list 命令
    list_parser = subparsers.add_parser('list', help='列出文章')
    list_parser.add_argument('--author', help='按作者筛选')
    list_parser.add_argument('--category', help='按分类筛选')
    list_parser.add_argument('--limit', type=int, default=20, help='返回数量')

    # search 命令
    search_parser = subparsers.add_parser('search', help='搜索文章')
    search_parser.add_argument('keyword', help='搜索关键词')
    search_parser.add_argument('--limit', type=int, default=20, help='返回数量')

    # stats 命令
    stats_parser = subparsers.add_parser('stats', help='统计信息')

    # export 命令
    export_parser = subparsers.add_parser('export', help='导出到 Excel')
    export_parser.add_argument('-o', '--output', required=True, help='输出文件路径')

    args = parser.parse_args()

    storage = ArticleStorage(args.db)

    if args.command == 'save':
        with open(args.file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        article_id, action = storage.save_article(data)
        print(f"文章已{ {'inserted': '新增', 'updated': '更新', 'unchanged': '无变化'}[action] }: ID={article_id}")

    elif args.command == 'get':
        article = storage.get_article(args.url)
        if article:
            print(json.dumps(article, ensure_ascii=False, indent=2))
        else:
            print("文章不存在")

    elif args.command == 'list':
        articles = storage.list_articles(
            author=args.author,
            category=args.category,
            limit=args.limit
        )
        print(f"共 {len(articles)} 篇文章:")
        for article in articles:
            print(f"  [{article.get('id')}] {article.get('title', '无标题')} - {article.get('author', '未知')}")

    elif args.command == 'search':
        articles = storage.search_articles(args.keyword, args.limit)
        print(f"搜索 '{args.keyword}' 找到 {len(articles)} 篇文章:")
        for article in articles:
            print(f"  [{article.get('id')}] {article.get('title', '无标题')}")

    elif args.command == 'stats':
        stats = storage.get_statistics()
        print(json.dumps(stats, ensure_ascii=False, indent=2))

    elif args.command == 'export':
        output_path = storage.export_to_excel(args.output)
        print(f"数据已导出: {output_path}")

    else:
        parser.print_help()


if __name__ == '__main__':
    main()
