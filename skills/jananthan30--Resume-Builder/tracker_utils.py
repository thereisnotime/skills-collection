"""
Job Application Tracker Utilities

This module provides functions to manage the Job_Application_Tracker.xlsx file.
It automatically creates/updates the tracker when new applications are generated.
"""

import os
import tempfile
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: pandas or openpyxl not installed. Excel tracking disabled.")


# Default tracker path
TRACKER_PATH = Path(__file__).parent / "Job_Application_Tracker.xlsx"
APPLICATIONS_DIR = Path(__file__).parent / "applications"


# Canonical column order for the tracker. New strategic columns appended at
# the end so existing trackers can be migrated by add_application().
TRACKER_COLUMNS = [
    'Company', 'Job Title', 'Application Date', 'Status',
    'Resume File', 'Cover Letter File', 'Job Description',
    'ATS Score', 'HR Score', 'Notes',
    'Interview Date', 'Follow Up Date', 'Response',
    # Strategic columns (Sprint 5) — used to learn from pipeline outcomes
    'Target Tier',                 # IC / Sr / Manager / AD / Director / Other
    'Fit Label',                   # MEETS / STRETCH / MISS (from job_fit_scorer)
    'Hard Reqs Missed',            # int — count of knockout flags at apply time
    'Referral Source',             # cold / alumni / recruiter / network / referral
    'Rejection Reason',            # taxonomy: no_response / auto_reject / screen_reject / interview_reject / offer_declined / withdrawn
    'Days To Response',            # int — auto-computed from dates
    'Interview Stages Reached',    # int — phone screen=1, hiring mgr=2, panel=3, onsite=4, offer=5
    # Evidence columns. These replaced ATS/HR as the meaningful record of how
    # well an application was supported; the keyword columns above are kept so
    # historical rows stay readable.
    'Evidence Match',              # % — qualification_evidence_score
    'Eligibility',                 # PASS / FAIL / UNVERIFIED
    'Requirements Evidenced',      # "5/8" — strongly evidenced over total
    'Must-Haves Missing',          # int — must-haves with no evidence found
    'Evidence Policy Version',     # which scoring policy produced the numbers
]


def format_excel_worksheet(worksheet, num_rows):
    """Apply formatting to the Excel worksheet."""
    # Column widths
    column_widths = {
        'A': 30,  # Company
        'B': 35,  # Job Title
        'C': 15,  # Application Date
        'D': 14,  # Status
        'E': 50,  # Resume File
        'F': 55,  # Cover Letter File
        'G': 20,  # Job Description
        'H': 11,  # ATS Score
        'I': 11,  # HR Score
        'J': 30,  # Notes
        'K': 15,  # Interview Date
        'L': 15,  # Follow Up Date
        'M': 20,  # Response
        # Strategic columns
        'N': 12,  # Target Tier
        'O': 12,  # Fit Label
        'P': 12,  # Hard Reqs Missed
        'Q': 16,  # Referral Source
        'R': 20,  # Rejection Reason
        'S': 12,  # Days To Response
        'T': 12,  # Interview Stages Reached
        # Evidence columns
        'U': 14,  # Evidence Match
        'V': 13,  # Eligibility
        'W': 21,  # Requirements Evidenced
        'X': 18,  # Must-Haves Missing
        'Y': 24,  # Evidence Policy Version
    }

    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    # Header formatting
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data row formatting
    for row in worksheet.iter_rows(min_row=2, max_row=num_rows + 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')


def _ensure_strategic_columns(df):
    """Ensure the dataframe has every column in TRACKER_COLUMNS, in order."""
    for col in TRACKER_COLUMNS:
        if col not in df.columns:
            df[col] = ''
    # Reorder to canonical order; drop any unexpected columns at the end
    extra = [c for c in df.columns if c not in TRACKER_COLUMNS]
    return df[TRACKER_COLUMNS + extra]


_BLANK_EVIDENCE_COLUMNS = {
    'Evidence Match': '',
    'Eligibility': '',
    'Requirements Evidenced': '',
    'Must-Haves Missing': '',
    'Evidence Policy Version': '',
}


def _evidence_columns(evidence: dict = None) -> dict:
    """Flatten an evidence summary into tracker columns.

    A failed or absent match leaves the columns blank rather than writing a
    zero, so an unscored application is never mistaken for a bad one.
    """
    if not isinstance(evidence, dict) or evidence.get("status") != "OK":
        return dict(_BLANK_EVIDENCE_COLUMNS)
    score = evidence.get("qualification_evidence_score")
    return {
        'Evidence Match': f"{score * 100:.1f}%" if score is not None else '',
        'Eligibility': evidence.get("eligibility", ''),
        'Requirements Evidenced': (
            f"{evidence.get('requirements_strong', 0)}/"
            f"{evidence.get('requirements_total', 0)}"),
        'Must-Haves Missing': len(evidence.get("missing_must_have_ids") or []),
        'Evidence Policy Version': evidence.get("scoring_policy_version", ''),
    }


def add_application(
    company: str,
    job_title: str,
    resume_file: str = "",
    cover_letter_file: str = "",
    jd_file: str = "job_description.txt",
    ats_score: float = None,
    hr_score: float = None,
    application_date: str = None,
    status: str = "Applied",
    notes: str = "",
    # Sprint 5 — strategic columns. All optional; default to blank/0/None.
    target_tier: str = "",
    fit_label: str = "",
    hard_reqs_missed: int = None,
    referral_source: str = "",
    rejection_reason: str = "",
    interview_stages_reached: int = None,
    # Evidence columns — pass the summarize() dict from an evidence match.
    evidence: dict = None,
):
    """
    Add a new application to the Job Application Tracker.

    Args:
        company: Company name
        job_title: Job title/position
        resume_file: Name of the resume file
        cover_letter_file: Name of the cover letter file
        jd_file: Name of the job description file
        ats_score: ATS score (0-100)
        hr_score: HR score (0-100)
        application_date: Date string (YYYY-MM-DD), defaults to today
        status: Application status (default: "Applied")
        notes: Any additional notes
        target_tier: IC / Sr / Manager / AD / Director (pipeline-mix tracking)
        fit_label: MEETS / STRETCH / MISS (from job_fit_scorer)
        hard_reqs_missed: count of knockout flags at apply time
        referral_source: cold / alumni / recruiter / network / referral
        rejection_reason: no_response / auto_reject / screen_reject /
            interview_reject / offer_declined / withdrawn
        interview_stages_reached: 0=applied, 1=phone screen, 2=hiring mgr,
            3=panel, 4=onsite, 5=offer
        evidence: ``evidence_engine.api.summarize()`` output. Records what the
            resume actually evidenced, which is the column worth learning from
            — ATS/HR remain only so older rows stay readable.

    Returns:
        bool: True if successful, False otherwise
    """
    if not EXCEL_AVAILABLE:
        print("Excel tracking not available. Install pandas and openpyxl.")
        return False
    # Default to today's date
    if application_date is None:
        application_date = datetime.now().strftime('%Y-%m-%d')

    # Prepare the new row
    new_row = {
        'Company': company,
        'Job Title': job_title,
        'Application Date': application_date,
        'Status': status,
        'Resume File': resume_file,
        'Cover Letter File': cover_letter_file,
        'Job Description': jd_file,
        'ATS Score': f"{ats_score:.1f}%" if ats_score else "",
        'HR Score': f"{hr_score:.1f}%" if hr_score else "",
        'Notes': notes,
        'Interview Date': '',
        'Follow Up Date': '',
        'Response': '',
        # Strategic columns
        'Target Tier': target_tier,
        'Fit Label': fit_label,
        'Hard Reqs Missed': hard_reqs_missed if hard_reqs_missed is not None else '',
        'Referral Source': referral_source,
        'Rejection Reason': rejection_reason,
        'Days To Response': '',  # filled by mark_response()
        'Interview Stages Reached': interview_stages_reached if interview_stages_reached is not None else '',
        **_evidence_columns(evidence),
    }

    temporary_path = None
    try:
        destination = Path(TRACKER_PATH)
        if destination.is_symlink():
            raise ValueError("tracker path must not be a symlink")
        destination.parent.mkdir(parents=True, exist_ok=True)
        existing_mode = (
            destination.stat().st_mode & 0o777 if destination.exists() else 0o600
        )
        # Check if tracker exists
        if destination.exists():
            # Load existing tracker
            df = pd.read_excel(destination, sheet_name='Applications')
            # Migrate: ensure new strategic columns exist
            df = _ensure_strategic_columns(df)

            # Check if this application already exists (same company + job title)
            mask = (df['Company'] == company) & (df['Job Title'] == job_title)
            if mask.any():
                # Update existing row
                idx = df[mask].index[0]
                for key, value in new_row.items():
                    if value != '' and value is not None:
                        df.at[idx, key] = value
                action_message = f"Updated existing application: {company} - {job_title}"
            else:
                # Add new row
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                action_message = f"Added new application: {company} - {job_title}"
        else:
            # Create new tracker
            df = pd.DataFrame([new_row], columns=TRACKER_COLUMNS)
            action_message = f"Created new tracker with application: {company} - {job_title}"

        # Sort by date (most recent first)
        df['Application Date'] = pd.to_datetime(df['Application Date'], errors='coerce')
        df = df.sort_values('Application Date', ascending=False)
        df['Application Date'] = df['Application Date'].dt.strftime('%Y-%m-%d')

        # Build and validate a sibling workbook before the single atomic
        # replacement.  Any formatter/serializer/verification exception leaves
        # a pre-existing tracker byte-for-byte untouched.
        handle = tempfile.NamedTemporaryFile(
            prefix=f'.{destination.name}.',
            suffix='.tmp.xlsx',
            dir=destination.parent,
            delete=False,
        )
        temporary_path = Path(handle.name)
        handle.close()
        with pd.ExcelWriter(temporary_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Applications', index=False)
            format_excel_worksheet(writer.sheets['Applications'], len(df))
        workbook = load_workbook(temporary_path, read_only=True)
        try:
            if 'Applications' not in workbook.sheetnames:
                raise ValueError("tracker workbook verification failed")
        finally:
            workbook.close()
        os.chmod(temporary_path, existing_mode)
        descriptor = os.open(temporary_path, os.O_RDONLY)
        try:
            os.fsync(descriptor)
        finally:
            os.close(descriptor)
        os.replace(temporary_path, destination)
        temporary_path = None

        print(action_message)
        print(f"Tracker saved: {destination}")
        return True

    except Exception as e:
        print(f"Error updating tracker: {e}")
        return False
    finally:
        if temporary_path is not None:
            try:
                temporary_path.unlink()
            except FileNotFoundError:
                pass


class TrackerUpdateError(RuntimeError):
    """A guarded tracker mutation did not complete successfully."""


def add_application_authorized(
    company: str,
    job_title: str,
    *,
    authorized_resume_path: str,
    receipt_path: str,
    expected_receipt_digest: str,
    config_path: str = "config.json",
    resume_file: str = "",
    cover_letter_file: str = "",
    jd_file: str = "job_description.txt",
    ats_score: float = None,
    hr_score: float = None,
    application_date: str = None,
    status: str = "Applied",
    notes: str = "",
    target_tier: str = "",
    fit_label: str = "",
    hard_reqs_missed: int = None,
    referral_source: str = "",
    rejection_reason: str = "",
    interview_stages_reached: int = None,
    evidence: dict = None,
) -> bool:
    """Revalidate a native-team receipt immediately before tracker mutation.

    A literal ``True`` from the legacy mutation is the only success state.
    False returns and exceptions propagate as ``TrackerUpdateError`` so callers
    cannot report a tracker update that did not happen.
    """

    from final_receipt_verifier import verify_final_receipt

    verify_final_receipt(
        resume_path=authorized_resume_path,
        receipt_path=receipt_path,
        expected_receipt_digest=expected_receipt_digest,
        config_path=config_path,
    )
    try:
        updated = add_application(
            company=company,
            job_title=job_title,
            resume_file=resume_file,
            cover_letter_file=cover_letter_file,
            jd_file=jd_file,
            ats_score=ats_score,
            hr_score=hr_score,
            application_date=application_date,
            status=status,
            notes=notes,
            target_tier=target_tier,
            fit_label=fit_label,
            hard_reqs_missed=hard_reqs_missed,
            referral_source=referral_source,
            rejection_reason=rejection_reason,
            interview_stages_reached=interview_stages_reached,
            evidence=evidence,
        )
    except Exception as exc:
        raise TrackerUpdateError('TRACKER_UPDATE_EXCEPTION') from exc
    if updated is not True:
        raise TrackerUpdateError('TRACKER_UPDATE_FAILED')
    return True


def mark_response(
    company: str,
    job_title: str,
    response_date: str = None,
    rejection_reason: str = "",
    interview_stages_reached: int = None,
    status: str = None,
):
    """
    Record an outcome for an existing application: rejection, screen advance,
    interview advance, offer, etc. Auto-computes Days To Response from
    Application Date if a response_date is supplied.

    Args:
        company: Company name (used to find the row)
        job_title: Job title (used to find the row)
        response_date: Date heard back (YYYY-MM-DD). Defaults to today.
        rejection_reason: taxonomy value (see add_application docstring)
        interview_stages_reached: 0–5 (see add_application docstring)
        status: optional status override (e.g., "Rejected", "Phone Screen",
            "Onsite", "Offer", "Withdrawn")
    """
    if not EXCEL_AVAILABLE or not TRACKER_PATH.exists():
        return False
    if response_date is None:
        response_date = datetime.now().strftime('%Y-%m-%d')

    try:
        df = pd.read_excel(TRACKER_PATH, sheet_name='Applications')
        df = _ensure_strategic_columns(df)
        mask = (df['Company'] == company) & (df['Job Title'] == job_title)
        if not mask.any():
            print(f"Application not found: {company} - {job_title}")
            return False
        idx = df[mask].index[0]

        # Days to response — from Application Date
        try:
            app_dt = pd.to_datetime(df.at[idx, 'Application Date'])
            resp_dt = pd.to_datetime(response_date)
            delta = (resp_dt - app_dt).days
            if delta >= 0:
                df.at[idx, 'Days To Response'] = int(delta)
        except Exception:
            pass

        df.at[idx, 'Response'] = response_date
        if rejection_reason:
            df.at[idx, 'Rejection Reason'] = rejection_reason
        if interview_stages_reached is not None:
            df.at[idx, 'Interview Stages Reached'] = int(interview_stages_reached)
        if status:
            df.at[idx, 'Status'] = status

        with pd.ExcelWriter(TRACKER_PATH, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Applications', index=False)
            format_excel_worksheet(writer.sheets['Applications'], len(df))
        print(f"Marked response for {company} - {job_title} ({response_date})")
        return True
    except Exception as e:
        print(f"Error marking response: {e}")
        return False


def pipeline_summary():
    """
    Print conversion summary by Target Tier and Fit Label so the user can see
    which segments of their pipeline actually convert.
    """
    if not EXCEL_AVAILABLE or not TRACKER_PATH.exists():
        print("No tracker found.")
        return None
    df = pd.read_excel(TRACKER_PATH, sheet_name='Applications')
    df = _ensure_strategic_columns(df)

    total = len(df)
    if total == 0:
        print("Tracker is empty.")
        return None

    def safe_int(s):
        try:
            return int(s) if s != '' else 0
        except Exception:
            return 0

    df['_stages'] = df['Interview Stages Reached'].apply(safe_int)
    df['_responded'] = df['Response'].apply(lambda v: bool(str(v).strip()))

    print(f'=== Pipeline summary — {total} applications ===\n')

    def _bucket(label_col):
        groups = df.groupby(label_col).agg(
            total=('Company', 'count'),
            responded=('_responded', 'sum'),
            interviewed=('_stages', lambda s: int((s >= 1).sum())),
            avg_stages=('_stages', 'mean'),
        )
        if groups.empty:
            return
        print(f'By {label_col}:')
        print(groups.to_string())
        print()

    _bucket('Target Tier')
    _bucket('Fit Label')
    _bucket('Referral Source')
    if 'Rejection Reason' in df.columns:
        rejs = df[df['Rejection Reason'].astype(str).str.strip() != '']
        if len(rejs) > 0:
            print('Rejection-reason taxonomy:')
            print(rejs['Rejection Reason'].value_counts().to_string())


def get_all_applications():
    """
    Get all applications from the tracker.

    Returns:
        pandas.DataFrame or None if tracker doesn't exist
    """
    if not EXCEL_AVAILABLE:
        return None

    if TRACKER_PATH.exists():
        return pd.read_excel(TRACKER_PATH, sheet_name='Applications')
    return None


def update_application_status(company: str, job_title: str, status: str, notes: str = None):
    """
    Update the status of an existing application.

    Args:
        company: Company name
        job_title: Job title
        status: New status (e.g., "Interview Scheduled", "Rejected", "Offer")
        notes: Optional notes to add
    """
    if not EXCEL_AVAILABLE or not TRACKER_PATH.exists():
        return False

    try:
        df = pd.read_excel(TRACKER_PATH, sheet_name='Applications')
        mask = (df['Company'] == company) & (df['Job Title'] == job_title)

        if mask.any():
            df.loc[mask, 'Status'] = status
            if notes:
                df.loc[mask, 'Notes'] = notes

            with pd.ExcelWriter(TRACKER_PATH, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Applications', index=False)
                format_excel_worksheet(writer.sheets['Applications'], len(df))

            print(f"Updated status for {company} - {job_title}: {status}")
            return True
        else:
            print(f"Application not found: {company} - {job_title}")
            return False

    except Exception as e:
        print(f"Error updating status: {e}")
        return False


def rebuild_tracker_from_folders():
    """
    Rebuild the entire tracker by scanning the applications folder.
    Useful if tracker gets out of sync or needs to be regenerated.
    """
    if not EXCEL_AVAILABLE:
        return False

    applications = []

    for folder in APPLICATIONS_DIR.iterdir():
        if folder.is_dir():
            folder_name = folder.name

            # Parse company and job title
            parts = folder_name.split(" - ", 1)
            if len(parts) == 2:
                company = parts[0].strip()
                job_title = parts[1].strip()
            else:
                company = folder_name
                job_title = ""

            # Get file info
            resume_file = ""
            cover_letter_file = ""
            jd_file = ""
            application_date = None

            for file in folder.iterdir():
                if file.is_file():
                    file_lower = file.name.lower()
                    if 'resume' in file_lower and file_lower.endswith('.docx'):
                        resume_file = file.name
                        application_date = datetime.fromtimestamp(file.stat().st_ctime)
                    elif 'cover' in file_lower and file_lower.endswith('.docx'):
                        cover_letter_file = file.name
                    elif file_lower.endswith('.txt'):
                        jd_file = file.name

            if application_date is None:
                for file in folder.iterdir():
                    if file.is_file():
                        application_date = datetime.fromtimestamp(file.stat().st_ctime)
                        break

            row = {col: '' for col in TRACKER_COLUMNS}
            row.update({
                'Company': company,
                'Job Title': job_title,
                'Application Date': application_date.strftime('%Y-%m-%d') if application_date else '',
                'Status': 'Applied',
                'Resume File': resume_file,
                'Cover Letter File': cover_letter_file,
                'Job Description': jd_file,
            })
            applications.append(row)

    # Sort and save
    applications.sort(key=lambda x: x['Application Date'], reverse=True)
    df = pd.DataFrame(applications, columns=TRACKER_COLUMNS)

    with pd.ExcelWriter(TRACKER_PATH, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Applications', index=False)
        format_excel_worksheet(writer.sheets['Applications'], len(df))

    print(f"Tracker rebuilt with {len(applications)} applications")
    return True


if __name__ == '__main__':
    # Test/example usage
    print("Job Application Tracker Utilities")
    print("=" * 50)

    if TRACKER_PATH.exists():
        df = get_all_applications()
        print(f"\nCurrent applications: {len(df)}")
        print(df[['Company', 'Job Title', 'Application Date', 'Status']].to_string())
    else:
        print("\nNo tracker found. Will be created when first application is added.")
