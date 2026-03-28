"""
Batch Job Search — Search multiple role types, triple-score, save top results to Excel.

Usage:
    python batch_job_search.py [--location "New York"] [--top 10]
"""

import sys
import io
import os
import re
from datetime import datetime
from typing import Any, Dict, List

# Fix Windows encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from dotenv import load_dotenv
load_dotenv()


def batch_search_and_score(
    resume_text: str,
    queries: List[str],
    location: str = "New York",
    top_n: int = 10,
    date_posted: str = "month",
) -> List[Dict[str, Any]]:
    """
    Search multiple job titles, merge, deduplicate, triple-score, rank.

    Returns top_n jobs sorted by combined score (Job Fit 40% + ATS 30% + HR 30%).
    """
    from job_discovery import (
        search_jsearch, search_adzuna, strip_html,
        _title_similarity, _jsearch_configured, _adzuna_configured,
    )

    # --- Step 1: Search all queries ---
    all_jobs: List[Dict[str, Any]] = []
    seen_ids: set = set()
    # Also deduplicate by normalized title+company
    seen_title_company: set = set()

    has_jsearch = _jsearch_configured()
    has_adzuna = _adzuna_configured()

    for i, query in enumerate(queries):
        print(f"  Searching [{i+1}/{len(queries)}]: {query}...", end=" ")
        count_before = len(all_jobs)

        if has_jsearch:
            for job in search_jsearch(query, location=location, date_posted=date_posted):
                job_key = f"{job['title'].lower().strip()}|{job['company'].lower().strip()}"
                if job["id"] not in seen_ids and job_key not in seen_title_company:
                    seen_ids.add(job["id"])
                    seen_title_company.add(job_key)
                    job["_search_query"] = query
                    all_jobs.append(job)

        if has_adzuna:
            for job in search_adzuna(query, location=location):
                job_key = f"{job['title'].lower().strip()}|{job['company'].lower().strip()}"
                if job["id"] not in seen_ids and job_key not in seen_title_company:
                    seen_ids.add(job["id"])
                    seen_title_company.add(job_key)
                    job["_search_query"] = query
                    all_jobs.append(job)

        added = len(all_jobs) - count_before
        print(f"{added} new jobs")

    print(f"\n  Total unique jobs found: {len(all_jobs)}")

    if not all_jobs:
        print("  No jobs found. Check API keys and try broader search terms.")
        return []

    # --- Step 2: Pre-filter by title similarity to ANY query ---
    for job in all_jobs:
        best_sim = max(_title_similarity(job["title"], q) for q in queries)
        job["_best_title_sim"] = best_sim

    # Keep top 25 by title similarity (wider pool for scoring)
    all_jobs.sort(key=lambda j: j["_best_title_sim"], reverse=True)
    candidates = all_jobs[:25]
    print(f"  Pre-filtered to top {len(candidates)} candidates by title relevance")

    # --- Step 3: Triple-score candidates ---
    import ats_scorer
    import hr_scorer
    from job_fit_scorer import calculate_job_fit

    scored_jobs = []
    print(f"\n  Scoring {len(candidates)} jobs (Job Fit + ATS + HR)...")

    for i, job in enumerate(candidates):
        desc = job.get("description", "")
        title = job.get("title", "Unknown")
        company = job.get("company", "Unknown")
        print(f"    [{i+1}/{len(candidates)}] {title[:50]} @ {company[:20]}...", end=" ")

        if not desc or len(desc) < 50:
            print("SKIP (no description)")
            continue

        entry = {
            "title": title,
            "company": company,
            "location": job.get("location", ""),
            "salary_min": job.get("salary_min"),
            "salary_max": job.get("salary_max"),
            "apply_url": job.get("url", ""),
            "listing_url": job.get("listing_url", job.get("url", "")),
            "source": job.get("source", ""),
            "posted_date": job.get("posted_date", ""),
            "employment_type": job.get("employment_type", ""),
            "is_remote": job.get("is_remote", False),
            "search_query": job.get("_search_query", ""),
            "description": desc,
        }

        # Job Fit
        try:
            fit_result = calculate_job_fit(resume_text, desc)
            entry["job_fit_score"] = round(fit_result.overall_score, 1)
            entry["job_fit_verdict"] = fit_result.recommendation
            knockout_flags = []
            if fit_result.knockouts and not fit_result.knockouts.passed:
                knockout_flags = [
                    k.requirement for k in fit_result.knockouts.knockouts
                ] if fit_result.knockouts.knockouts else []
            entry["knockouts"] = "; ".join(knockout_flags) if knockout_flags else ""
        except Exception:
            entry["job_fit_score"] = 50.0
            entry["job_fit_verdict"] = "ERROR"
            entry["knockouts"] = ""

        # ATS
        try:
            ats_result = ats_scorer.calculate_ats_score(resume_text, desc)
            entry["ats_score"] = round(ats_result.get("total_score", 0), 1)
            entry["matched_keywords"] = ", ".join(ats_result.get("matched_keywords", [])[:10])
            entry["missing_keywords"] = ", ".join(ats_result.get("missing_keywords", [])[:10])
        except Exception:
            entry["ats_score"] = 0.0
            entry["matched_keywords"] = ""
            entry["missing_keywords"] = ""

        # HR
        try:
            hr_result = hr_scorer.calculate_hr_score_from_text(resume_text, desc)
            hr_dict = hr_scorer.result_to_dict(hr_result)
            entry["hr_score"] = round(hr_dict.get("overall_score", 0), 1)
            entry["hr_recommendation"] = hr_dict.get("recommendation", "")
        except Exception:
            entry["hr_score"] = 0.0
            entry["hr_recommendation"] = ""

        # Combined score
        entry["combined_score"] = round(
            entry["job_fit_score"] * 0.4
            + entry["ats_score"] * 0.3
            + entry["hr_score"] * 0.3, 1
        )

        print(f"FIT:{entry['job_fit_score']:.0f} ATS:{entry['ats_score']:.0f} HR:{entry['hr_score']:.0f}")
        scored_jobs.append(entry)

    # --- Step 4: Rank by combined score, take top N ---
    scored_jobs.sort(key=lambda j: j["combined_score"], reverse=True)
    top_jobs = scored_jobs[:top_n]

    # Assign ranks
    for i, job in enumerate(top_jobs, 1):
        job["rank"] = i

    return top_jobs


def save_to_excel(jobs: List[Dict[str, Any]], output_path: str):
    """Save scored jobs to Excel with formatting."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top Job Matches"

    # Headers
    headers = [
        "Rank", "Combined", "Job Fit", "ATS", "HR", "Verdict",
        "Job Title", "Company", "Location", "Salary Range",
        "Apply URL", "Source", "Posted", "Search Query",
        "Matched Keywords", "Missing Keywords", "Knockouts",
        "HR Recommendation", "Status", "Notes"
    ]

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Score color coding
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def score_fill(score):
        if score >= 75:
            return green_fill
        elif score >= 60:
            return yellow_fill
        return red_fill

    # Data rows
    for row_idx, job in enumerate(jobs, 2):
        sal_min = job.get("salary_min")
        sal_max = job.get("salary_max")
        salary = ""
        if sal_min and sal_max:
            salary = f"${sal_min:,.0f} - ${sal_max:,.0f}"
        elif sal_min:
            salary = f"${sal_min:,.0f}+"

        row_data = [
            job.get("rank", ""),
            job.get("combined_score", 0),
            job.get("job_fit_score", 0),
            job.get("ats_score", 0),
            job.get("hr_score", 0),
            job.get("job_fit_verdict", ""),
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            salary,
            job.get("apply_url", ""),
            job.get("source", ""),
            job.get("posted_date", ""),
            job.get("search_query", ""),
            job.get("matched_keywords", ""),
            job.get("missing_keywords", ""),
            job.get("knockouts", ""),
            job.get("hr_recommendation", ""),
            "",  # Status (empty for user)
            "",  # Notes (empty for user)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col >= 7))

        # Color-code score cells
        for col_idx in [2, 3, 4, 5]:  # Combined, Fit, ATS, HR
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = score_fill(cell.value or 0)
            cell.number_format = '0.0'

        # Make Apply URL a hyperlink
        url_cell = ws.cell(row=row_idx, column=11)
        if url_cell.value:
            url_cell.hyperlink = url_cell.value
            url_cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    # Column widths
    col_widths = {
        1: 6, 2: 10, 3: 8, 4: 8, 5: 8, 6: 14,
        7: 45, 8: 25, 9: 20, 10: 22,
        11: 40, 12: 18, 13: 12, 14: 25,
        15: 30, 16: 30, 17: 25,
        18: 15, 19: 12, 20: 20,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:T{len(jobs)+1}"

    # Save JD text to a second sheet (for resume generation later)
    ws_jd = wb.create_sheet("Job Descriptions")
    ws_jd.cell(row=1, column=1, value="Rank").font = Font(bold=True)
    ws_jd.cell(row=1, column=2, value="Title").font = Font(bold=True)
    ws_jd.cell(row=1, column=3, value="Company").font = Font(bold=True)
    ws_jd.cell(row=1, column=4, value="Job Description").font = Font(bold=True)
    ws_jd.column_dimensions['A'].width = 6
    ws_jd.column_dimensions['B'].width = 40
    ws_jd.column_dimensions['C'].width = 25
    ws_jd.column_dimensions['D'].width = 100

    for row_idx, job in enumerate(jobs, 2):
        ws_jd.cell(row=row_idx, column=1, value=job.get("rank", ""))
        ws_jd.cell(row=row_idx, column=2, value=job.get("title", ""))
        ws_jd.cell(row=row_idx, column=3, value=job.get("company", ""))
        desc_cell = ws_jd.cell(row=row_idx, column=4, value=job.get("description", ""))
        desc_cell.alignment = Alignment(wrap_text=True)

    wb.save(output_path)
    print(f"\n  Saved to: {output_path}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Batch job search with triple scoring")
    parser.add_argument("--location", default="New York", help="Location filter")
    parser.add_argument("--top", type=int, default=10, help="Top N results")
    parser.add_argument("--output", default="", help="Output Excel path")
    parser.add_argument("--queries", nargs="*", default=None, help="Custom search queries")
    args = parser.parse_args()

    # Default queries for MD clinical research professional
    queries = args.queries or [
        "Medical Monitor MD",
        "Clinical Research Scientist Associate Director",
        "Drug Safety Physician",
        "Senior Medical Writer pharma",
        "Associate Director Clinical Operations",
    ]

    # Output path
    today = datetime.now().strftime("%Y-%m-%d")
    output_path = args.output or f"Job_Search_Results_{today}.xlsx"

    print(f"\n{'='*70}")
    print(f"  BATCH JOB SEARCH — Triple Scorer (Job Fit + ATS + HR)")
    print(f"{'='*70}")
    print(f"  Location: {args.location}")
    print(f"  Queries:  {len(queries)}")
    for q in queries:
        print(f"    - {q}")
    print(f"  Top N:    {args.top}")
    print(f"{'='*70}\n")

    # Load resume
    import json
    config_path = os.path.join(os.path.dirname(__file__), "config.json")
    with open(config_path, 'r') as f:
        config = json.load(f)
    resume_path = config.get("master_resume_path", "master_resume.md")
    with open(resume_path, 'r', encoding='utf-8') as f:
        resume_text = f.read()

    print("  Step 1: Searching APIs...")
    top_jobs = batch_search_and_score(
        resume_text=resume_text,
        queries=queries,
        location=args.location,
        top_n=args.top,
    )

    if not top_jobs:
        print("\n  No scored jobs to save.")
        return

    print(f"\n  Step 2: Saving top {len(top_jobs)} to Excel...")
    save_to_excel(top_jobs, output_path)

    # Print summary table
    print(f"\n{'='*70}")
    print(f"  TOP {len(top_jobs)} JOBS — RANKED BY COMBINED SCORE")
    print(f"{'='*70}")
    print(f"{'#':>2} {'CMB':>5} {'FIT':>4} {'ATS':>5} {'HR':>5} {'Verdict':<12} {'Title':<40} {'Company':<20}")
    print("-" * 105)
    for j in top_jobs:
        print(
            f"{j['rank']:>2} "
            f"{j['combined_score']:>5.1f} "
            f"{j['job_fit_score']:>4.0f} "
            f"{j['ats_score']:>5.1f} "
            f"{j['hr_score']:>5.1f} "
            f"{j.get('job_fit_verdict',''):<12} "
            f"{j['title'][:39]:<40} "
            f"{j['company'][:19]:<20}"
        )
    print(f"\n  File: {output_path}")
    print(f"  Next: Open Excel, pick jobs to apply to, then run /resume for each")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
